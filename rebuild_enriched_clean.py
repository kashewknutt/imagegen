#!/usr/bin/env python3
"""
Rebuild outputs/stock_enriched.xlsx cleanly from scratch.

Columns:
  - All named (non-None) headers from Stock.xlsx (Total sheet)
  - description   — Shopify product descriptionHtml
  - thumbnailImageName — Shopify title + " " + SKU
  - thumbnailImage     — embedded latest prompt2_vX image
  - thumbnailImagePath — relative path like "SKU/prompt2_vX.jpg"

Rows come from SKU folders found in outputs/. Each row's stock data is
read from Stock.xlsx (matched by SKU column). Title and description are
fetched live from Shopify using credentials in root .env.

Usage:
  python rebuild_enriched_clean.py
  python rebuild_enriched_clean.py --output outputs/stock_enriched_v2.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env", override=False)

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from src.media_workspace import resolve_thumbnail_path, thumbnail_relative_path
from src.shopify_env import ensure_shopify_connection

OUTPUTS_DIR = ROOT / "outputs"
STOCK_PATH = ROOT / "Stock.xlsx"
OUTPUT_PATH = OUTPUTS_DIR / "stock_enriched.xlsx"

THUMB_MAX_PX = 120
THUMB_ROW_H = 95
THUMB_COL_W = 18


def _sku_folders(outputs_dir: Path) -> list[str]:
    return sorted(
        p.name
        for p in outputs_dir.iterdir()
        if p.is_dir() and p.name and p.name[0].isupper()
    )


def _read_stock(stock_path: Path) -> tuple[list[str], dict[str, dict]]:
    """Return (headers, {sku: row_dict}) from the Total sheet."""
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    ws = wb["Total"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    raw_headers = list(rows[0]) if rows else []
    seen: dict[str, int] = {}
    headers: list[str] = []
    for h in raw_headers:
        if h is None:
            continue
        name = str(h)
        count = seen.get(name, 0)
        seen[name] = count + 1
        headers.append(name if count == 0 else f"{name}_{count + 1}")

    by_sku: dict[str, dict] = {}
    sku_idx = None
    for i, h in enumerate(raw_headers):
        if str(h or "").strip().upper() == "SKU":
            sku_idx = i
            break

    for raw in rows[1:]:
        if sku_idx is None:
            break
        sku = str(raw[sku_idx] if sku_idx < len(raw) else "").strip()
        if not sku:
            continue
        row_dict: dict[str, object] = {}
        col_seen: dict[str, int] = {}
        for i, h in enumerate(raw_headers):
            if h is None:
                continue
            name = str(h)
            cnt = col_seen.get(name, 0)
            col_seen[name] = cnt + 1
            key = name if cnt == 0 else f"{name}_{cnt + 1}"
            row_dict[key] = raw[i] if i < len(raw) else None
        by_sku[sku] = row_dict

    return headers, by_sku


def _fetch_shopify_data(outputs_dir: Path) -> dict[str, dict]:
    """Return {sku: {title, description}} from Shopify."""
    print("Connecting to Shopify…")
    conn = ensure_shopify_connection(outputs_dir, env_path=ROOT / ".env")
    if not conn.connected or not conn.client:
        print(f"  WARNING: Shopify not connected — {conn.error or conn.status_label}")
        print("  Title/description will be empty. Set SHOPIFY_CLIENT_SECRET in .env.")
        return {}

    print(f"  Connected: {conn.shop_domain}")
    client = conn.client

    by_sku: dict[str, dict] = {}
    after = None
    page = 0
    while True:
        page += 1
        result = client.list_products(first=250, after=after)
        for prod in result.get("products") or []:
            title = str(prod.get("title") or "").strip()
            description = str(prod.get("description_html") or "").strip()
            for sku in prod.get("skus") or []:
                if sku:
                    by_sku[sku] = {"title": title, "description": description}
        page_info = result.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        after = page_info.get("endCursor")

    print(f"  Fetched {len(by_sku)} SKU→product mappings ({page} page(s))")
    return by_sku


def _embed_thumbnails(ws, *, outputs_dir: Path, sku_col: int, img_col: int, start_row: int = 2) -> int:
    col_letter = get_column_letter(img_col)
    ws.column_dimensions[col_letter].width = THUMB_COL_W
    embedded = 0
    for row_idx in range(start_row, ws.max_row + 1):
        sku = str(ws.cell(row=row_idx, column=sku_col).value or "").strip()
        if not sku:
            continue
        path = resolve_thumbnail_path(outputs_dir=outputs_dir, sku=sku)
        if path is None:
            continue
        try:
            img = XLImage(str(path))
            w, h = float(img.width), float(img.height)
            if w > 0 and h > 0:
                scale = min(THUMB_MAX_PX / w, THUMB_MAX_PX / h, 1.0)
                img.width = int(w * scale)
                img.height = int(h * scale)
            ws.add_image(img, f"{col_letter}{row_idx}")
            ws.row_dimensions[row_idx].height = THUMB_ROW_H
            ws.cell(row=row_idx, column=img_col).value = None
            embedded += 1
        except Exception as exc:
            print(f"  WARN: could not embed image for {sku}: {exc}")
    return embedded


def rebuild(
    *,
    stock_path: Path = STOCK_PATH,
    outputs_dir: Path = OUTPUTS_DIR,
    output_path: Path = OUTPUT_PATH,
) -> int:
    print(f"Reading Stock.xlsx headers from {stock_path} …")
    stock_headers, stock_by_sku = _read_stock(stock_path)
    print(f"  {len(stock_headers)} named columns, {len(stock_by_sku)} stock SKUs")

    shopify_by_sku = _fetch_shopify_data(outputs_dir)

    sku_folders = _sku_folders(outputs_dir)
    print(f"Found {len(sku_folders)} SKU folders in {outputs_dir}")

    extra_cols = ["description", "thumbnailImageName", "thumbnailImage", "thumbnailImagePath"]
    headers = stock_headers + extra_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Total"
    ws.append(headers)

    sku_col_idx = headers.index("SKU") + 1
    img_col_idx = headers.index("thumbnailImage") + 1

    written = 0
    skipped = 0
    for sku in sku_folders:
        stock_row = stock_by_sku.get(sku, {})
        shopify = shopify_by_sku.get(sku, {})

        shopify_title = shopify.get("title", "")
        shopify_desc = shopify.get("description", "")

        thumb_rel = thumbnail_relative_path(outputs_dir=outputs_dir, sku=sku)

        base = []
        for h in stock_headers:
            val = stock_row.get(h)
            if h == "title" and shopify_title:
                val = shopify_title
            base.append(val if val is not None else "")

        thumbnail_name = f"{shopify_title} {sku}".strip() if (shopify_title or sku) else ""

        extras = [
            shopify_desc,
            thumbnail_name,
            "",            # placeholder — image embedded after
            thumb_rel,
        ]
        ws.append(base + extras)
        written += 1

    print(f"Written {written} rows ({skipped} skipped). Embedding thumbnails…")
    embedded = _embed_thumbnails(ws, outputs_dir=outputs_dir, sku_col=sku_col_idx, img_col=img_col_idx)
    print(f"  Embedded {embedded} thumbnail image(s).")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved → {output_path}")
    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild stock_enriched.xlsx cleanly")
    parser.add_argument("--stock", type=Path, default=STOCK_PATH)
    parser.add_argument("--outputs-dir", type=Path, default=OUTPUTS_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    args = parser.parse_args()
    count = rebuild(stock_path=args.stock, outputs_dir=args.outputs_dir, output_path=args.output)
    print(f"Done — {count} rows.")


if __name__ == "__main__":
    main()
