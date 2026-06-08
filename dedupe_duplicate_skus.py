#!/usr/bin/env python3
"""
Find and remove duplicate SKU entries across Stock.xlsx and Shopify.

Shopify Review lists live Shopify products; duplicate SKUs there mean multiple
products were created with the same variant SKU (not sheet rows).

Usage:
  python dedupe_duplicate_skus.py --audit
  python dedupe_duplicate_skus.py --fix-stock
  python dedupe_duplicate_skus.py --archive-shopify-duplicates
"""
from __future__ import annotations

import argparse
import logging
from collections import Counter
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from src.config import load_config
from src.review_store import ReviewStore
from src.shopify_product_dedup import (
    fetch_all_shopify_products,
    group_shopify_products_by_sku,
    split_canonical_and_duplicates,
)
from src.xlsx_ingest import iter_rows

log = logging.getLogger("dedupe_duplicate_skus")


def audit_stock(stock_path: Path) -> list[tuple[str, int]]:
    rows = iter_rows(stock_path, ["Total"])
    counts = Counter(str(r.values.get("SKU") or "").strip() for r in rows if str(r.values.get("SKU") or "").strip())
    return [(sku, n) for sku, n in sorted(counts.items()) if n > 1]


def fix_stock_duplicates(stock_path: Path) -> int:
    wb = load_workbook(stock_path)
    ws = wb["Total"] if "Total" in wb.sheetnames else wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        sku_i = headers.index("SKU")
    except ValueError:
        raise RuntimeError("Stock sheet missing SKU column")

    seen: set[str] = set()
    removed = 0
    for row_idx in range(ws.max_row, 1, -1):
        sku = str(ws.cell(row=row_idx, column=sku_i + 1).value or "").strip()
        if not sku:
            continue
        if sku in seen:
            ws.delete_rows(row_idx, 1)
            removed += 1
            log.info("Removed duplicate Stock row for SKU %s (row %d)", sku, row_idx)
        else:
            seen.add(sku)
    if removed:
        wb.save(stock_path)
    wb.close()
    return removed


def audit_products_export(path: Path) -> list[tuple[str, int]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    sku_i = headers.index("Variant SKU")
    counts: Counter[str] = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_i] or "").strip()
        if sku:
            counts[sku] += 1
    wb.close()
    return [(sku, n) for sku, n in sorted(counts.items()) if n > 1]


def main() -> int:
    load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")

    parser = argparse.ArgumentParser(description="Audit and dedupe duplicate SKUs")
    parser.add_argument("--audit", action="store_true", help="Report duplicates only")
    parser.add_argument("--fix-stock", action="store_true", help="Remove duplicate SKU rows from Stock.xlsx (keep first)")
    parser.add_argument(
        "--archive-shopify-duplicates",
        action="store_true",
        help="Archive duplicate Shopify products, keeping canonical product per SKU",
    )
    args = parser.parse_args()
    if not (args.audit or args.fix_stock or args.archive_shopify_duplicates):
        args.audit = True

    cfg = load_config()
    root = Path(__file__).resolve().parent
    stock_path = cfg.xlsx_path
    products_path = root / "products_export_1.xlsx"

    stock_dups = audit_stock(stock_path)
    export_dups = audit_products_export(products_path) if products_path.is_file() else []
    log.info("Stock.xlsx duplicate SKUs: %d", len(stock_dups))
    for sku, n in stock_dups:
        log.info("  Stock %s — %d rows", sku, n)
    log.info("products_export duplicate Variant SKUs: %d", len(export_dups))
    for sku, n in export_dups[:20]:
        log.info("  export %s — %d rows", sku, n)
    if len(export_dups) > 20:
        log.info("  ... and %d more", len(export_dups) - 20)

    shopify_dupes: list[tuple[str, dict, dict]] = []
    try:
        from dedupe_titles_and_upload import _load_shopify_client

        client = _load_shopify_client(cfg.outputs_dir)
        review_store = ReviewStore(cfg.outputs_dir / "review_state.json")
        products = fetch_all_shopify_products(client)
        groups = group_shopify_products_by_sku(products)
        _canonical, shopify_dupes = split_canonical_and_duplicates(groups, review_store=review_store)
        multi = {k: v for k, v in groups.items() if len(v) > 1}
        log.info("Shopify products: %d | unique SKUs: %d | SKUs on multiple products: %d", len(products), len(groups), len(multi))
        for sku in sorted(multi):
            log.info("  Shopify %s — %d products", sku, len(multi[sku]))
            for p in multi[sku]:
                log.info("    %s | %s", str(p.get("handle") or ""), str(p.get("title") or "")[:60])
    except Exception as e:
        log.warning("Shopify audit skipped: %s", e)
        client = None

    if args.fix_stock:
        n = fix_stock_duplicates(stock_path)
        log.info("Removed %d duplicate row(s) from %s", n, stock_path)

    if args.archive_shopify_duplicates:
        if client is None:
            log.error("Cannot archive Shopify duplicates without API connection.")
            return 1
        archived = 0
        for sku, keep, dup in shopify_dupes:
            dup_id = str(dup.get("id") or "")
            keep_id = str(keep.get("id") or "")
            if not dup_id:
                continue
            try:
                client.product_update_status(product_id=dup_id, status="ARCHIVED")
                log.info("Archived duplicate %s | keep=%s | archived=%s", sku, keep_id[-8:], dup_id[-8:])
                archived += 1
            except Exception as e:
                log.error("Failed to archive %s (%s): %s", sku, dup_id, e)
        log.info("Archived %d duplicate Shopify product(s)", archived)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
