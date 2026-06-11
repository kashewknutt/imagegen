"""Gallery catalog and row builder for drive_review batch final pass."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from src.media_workspace import SkuMediaIndex, index_sku_media
from src.shopify_media_sync import media_paths_for_sku
from src.text_format import title_case_category

OPTIONAL_ENRICHED_COLUMNS = frozenset({
    "thumbnailImage",
    "thumbnail image path",
    "thumbnailImageName",
    "productDescription",
    "hashtag/keyword",
})

MINIMUM_REQUIRED_COLUMNS = frozenset({"SKU", "category", "productName"})


def load_enriched_index(path: Path) -> tuple[dict[str, dict[str, Any]], list[str]]:
    """Load enriched stock xlsx; returns (sku -> row dict, header list)."""
    from openpyxl import load_workbook

    if not path.is_file():
        return {}, []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        sku = str(vals.get("SKU") or "").strip()
        if sku and sku not in out:
            out[sku] = vals
    wb.close()
    return out, headers


def required_stock_columns(
    headers: list[str],
    enriched_index: dict[str, dict[str, Any]],
    *,
    fill_ratio: float = 0.75,
) -> list[str]:
    """
    Columns that must be non-empty per SKU for gallery inclusion.
    Uses core sheet columns present in >= fill_ratio of rows (not every sparse Excel col).
    """
    n = len(enriched_index) or 1
    threshold = max(1, int(n * fill_ratio))
    seen: set[str] = set()
    out: list[str] = []
    for h in headers:
        if not h or h in OPTIONAL_ENRICHED_COLUMNS or h in seen:
            continue
        if h in MINIMUM_REQUIRED_COLUMNS:
            seen.add(h)
            out.append(h)
            continue
        filled = sum(1 for row in enriched_index.values() if _cell_filled(row.get(h)))
        if filled >= threshold:
            seen.add(h)
            out.append(h)
    for m in MINIMUM_REQUIRED_COLUMNS:
        if m in headers and m not in seen:
            out.insert(0, m)
    return out


def _cell_filled(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    return s not in {"", "None", "nan"}


def _strip_html(text: str) -> str:
    return re.sub(r"<[^>]+>", " ", text or "").replace("&nbsp;", " ").strip()


def _shopify_media_counts(shop_prod: dict[str, Any] | None) -> tuple[int, int]:
    if not shop_prod:
        return 0, 0
    images = videos = 0
    for m in shop_prod.get("media") or []:
        ct = str(m.get("content_type") or "IMAGE").upper()
        if ct in {"VIDEO", "EXTERNAL_VIDEO"}:
            videos += 1
        else:
            images += 1
    return images, videos


def is_complete_sku(
    sku: str,
    *,
    enriched_row: dict[str, Any] | None,
    media_idx: SkuMediaIndex,
    required_cols: list[str],
    cfg,
    review_store=None,
) -> bool:
    if not enriched_row:
        return False
    for col in required_cols:
        if not _cell_filled(enriched_row.get(col)):
            return False
    if not media_idx.prompt1_versions or not media_idx.prompt2_versions:
        return False
    paths = media_paths_for_sku(cfg.base if hasattr(cfg, "base") else cfg, sku, review_store=review_store)
    if not paths.get("raw"):
        return False
    if not paths.get("videos") and not media_idx.videos:
        return False
    return True


def list_gallery_skus(
    cfg,
    enriched_index: dict[str, dict[str, Any]],
    *,
    required_cols: list[str],
    review_store=None,
) -> list[str]:
    outputs_dir = cfg.outputs_dir
    out: list[str] = []
    for sku in sorted(enriched_index.keys()):
        row = enriched_index[sku]
        media_idx = index_sku_media(outputs_dir=outputs_dir, sku=sku)
        if is_complete_sku(
            sku,
            enriched_row=row,
            media_idx=media_idx,
            required_cols=required_cols,
            cfg=cfg,
            review_store=review_store,
        ):
            out.append(sku)
    return out


def distinct_categories(
    enriched_index: dict[str, dict[str, Any]],
    skus: list[str],
) -> list[str]:
    cats: set[str] = set()
    for sku in skus:
        cat = title_case_category(str((enriched_index.get(sku) or {}).get("category") or ""))
        if cat:
            cats.add(cat)
    return sorted(cats)


def filter_by_category(
    skus: list[str],
    enriched_index: dict[str, dict[str, Any]],
    category: str,
) -> list[str]:
    if not category or category == "ALL":
        return list(skus)
    want = title_case_category(category)
    return [
        sku
        for sku in skus
        if title_case_category(str((enriched_index.get(sku) or {}).get("category") or "")) == want
    ]


def paginate(skus: list[str], *, page: int, page_size: int) -> tuple[list[str], int, int]:
    """Return (page_skus, total_pages, total_count). page is 1-based."""
    total = len(skus)
    if total == 0 or page_size <= 0:
        return [], 0, total
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    return skus[start : start + page_size], total_pages, total


@dataclass
class GalleryRow:
    sku: str
    title: str
    category: str
    description: str
    review_status: str
    in_sheet: bool
    on_local: bool
    on_drive: bool
    on_shopify: bool
    save_ready: bool
    local_raw: int
    local_videos: int
    local_has_p1: bool
    local_has_p2: bool
    drive_raw: int
    drive_videos: int
    drive_has_p1: bool
    drive_has_p2: bool
    shopify_images: int
    shopify_videos: int
    product_id: str
    readiness: dict[str, Any] = field(default_factory=dict)
    enriched_row: dict[str, Any] = field(default_factory=dict)


def display_title(
    *,
    enriched_row: dict[str, Any],
    review_rec: dict[str, Any],
    shop_prod: dict[str, Any] | None,
) -> str:
    for src in (
        str(enriched_row.get("productName") or "").strip(),
        str(review_rec.get("title") or "").strip(),
        str((shop_prod or {}).get("title") or "").strip(),
    ):
        if src:
            return src
    return ""


def display_category(
    *,
    enriched_row: dict[str, Any],
    review_rec: dict[str, Any],
) -> str:
    return title_case_category(
        str(enriched_row.get("category") or review_rec.get("category") or review_rec.get("product_type") or "")
    )


def display_description(
    *,
    enriched_row: dict[str, Any],
    review_rec: dict[str, Any],
    shop_prod: dict[str, Any] | None,
) -> str:
    for src in (
        str(enriched_row.get("productDescription") or "").strip(),
        str(review_rec.get("description") or "").strip(),
        _strip_html(str((shop_prod or {}).get("description_html") or "")),
    ):
        if src:
            return src
    return ""


def local_workspace_needs_drive_pull(
    cfg,
    sku: str,
    *,
    ref_paths: list[Path] | None = None,
) -> bool:
    """Skip Drive download when local prompts + raw refs are already present."""
    media_idx = index_sku_media(outputs_dir=cfg.outputs_dir, sku=sku)
    if not media_idx.prompt1_versions or not media_idx.prompt2_versions:
        return True
    if media_idx.raw_images:
        return False
    if ref_paths:
        return False
    return True


def gallery_readiness_light(
    *,
    media_idx: SkuMediaIndex,
    drive_meta: dict[str, Any] | None,
    shop_prod: dict[str, Any] | None,
) -> dict[str, Any]:
    """Local + cached Drive metadata only — no per-SKU Drive API calls."""
    drive_meta = drive_meta or {}
    shop_images, shop_videos = _shopify_media_counts(shop_prod)
    local = {
        "raw_count": len(media_idx.raw_images),
        "video_count": len(media_idx.videos),
        "has_prompt1": bool(media_idx.prompt1_versions),
        "has_prompt2": bool(media_idx.prompt2_versions),
        "has_raw": bool(media_idx.raw_images),
        "has_videos": bool(media_idx.videos),
    }
    blocking: list[str] = []
    if not local["has_prompt1"] or not local["has_prompt2"]:
        blocking.append("missing_prompts")
    if not local["has_raw"]:
        blocking.append("missing_raw")
    return {
        "local": local,
        "drive": drive_meta,
        "shopify_before": {
            "connected": shop_prod is not None,
            "image_count": shop_images,
            "video_count": shop_videos,
        },
        "blocking": blocking,
        "warnings": [],
        "ready": not blocking,
    }


def build_gallery_row(
    sku: str,
    *,
    cfg,
    enriched_row: dict[str, Any],
    media_idx: SkuMediaIndex,
    drive_meta: dict[str, Any] | None,
    shop_prod: dict[str, Any] | None,
    review_rec: dict[str, Any],
    readiness: dict[str, Any] | None = None,
    drive_folders: set[str] | None = None,
) -> GalleryRow:
    drive_meta = drive_meta or {}
    shop_images, shop_videos = _shopify_media_counts(shop_prod)
    loc = (readiness or {}).get("local") or {}
    return GalleryRow(
        sku=sku,
        title=display_title(enriched_row=enriched_row, review_rec=review_rec, shop_prod=shop_prod),
        category=display_category(enriched_row=enriched_row, review_rec=review_rec),
        description=display_description(enriched_row=enriched_row, review_rec=review_rec, shop_prod=shop_prod),
        review_status=str(review_rec.get("review_status") or "pending_review"),
        in_sheet=bool(enriched_row),
        on_local=media_idx.workspace_dir.is_dir(),
        on_drive=sku in (drive_folders or set()) or bool(drive_meta.get("folder_id")),
        on_shopify=shop_prod is not None,
        save_ready=bool((readiness or {}).get("ready")),
        local_raw=int(loc.get("raw_count") or len(media_idx.raw_images)),
        local_videos=int(loc.get("video_count") or len(media_idx.videos)),
        local_has_p1=bool(media_idx.prompt1_versions),
        local_has_p2=bool(media_idx.prompt2_versions),
        drive_raw=int(drive_meta.get("raw_count") or 0),
        drive_videos=int(drive_meta.get("video_count") or 0),
        drive_has_p1=bool(drive_meta.get("has_prompt1")),
        drive_has_p2=bool(drive_meta.get("has_prompt2")),
        shopify_images=shop_images,
        shopify_videos=shop_videos,
        product_id=str(review_rec.get("product_id") or (shop_prod or {}).get("id") or ""),
        readiness=readiness or {},
        enriched_row=enriched_row,
    )


def scan_drive_metadata_for_skus(
    sync,
    skus: list[str],
    *,
    drive_folders: dict[str, str] | None = None,
) -> dict[str, dict[str, Any]]:
    """Metadata-only Drive scan for the current gallery page."""
    folders = drive_folders if drive_folders is not None else sync.list_sku_folders(refresh=False)
    out: dict[str, dict[str, Any]] = {}
    for sku in skus:
        if sku not in folders:
            out[sku] = {"sku": sku, "has_prompt1": False, "has_prompt2": False, "raw_count": 0, "video_count": 0}
            continue
        try:
            out[sku] = sync.scan_sku_metadata(sku, folder_id=folders[sku])
        except Exception as e:
            out[sku] = {
                "sku": sku,
                "error": str(e),
                "has_prompt1": False,
                "has_prompt2": False,
                "raw_count": 0,
                "video_count": 0,
            }
    return out
