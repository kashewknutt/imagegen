"""Sync local latest prompt1/prompt2 versions to Drive, Shopify, and enriched XLSX."""
from __future__ import annotations

import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from build_stock_export import build_export
from openpyxl import load_workbook

from src.drive_outputs_sync import DriveOutputsSync
from src.drive_review_config import DriveReviewConfig
from src.drive_stock_sync import replace_drive_spreadsheet, resolve_stock_path
from src.media_workspace import (
    list_prompt_versions,
    sku_workspace_dir,
    thumbnail_relative_path,
)
from src.review_store import ReviewStore
from src.shopify_media_sync import replace_prompt_images_on_product
from src.typo_sku_cleanup import list_output_sku_dirs

PROMPT_VERSION_FILE_RE = re.compile(r"^prompt([12])_v(\d+)\.", re.IGNORECASE)


def _now_utc() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def local_prompt_state(outputs_dir: Path, sku: str) -> dict[str, dict[str, Any]]:
    """Latest prompt1/prompt2 version + filename from local workspace."""
    sku_dir = sku_workspace_dir(outputs_dir, sku)
    out: dict[str, dict[str, Any]] = {}
    for slot in ("prompt1", "prompt2"):
        versions = list_prompt_versions(sku_dir, slot)
        if not versions:
            continue
        ver, path = versions[-1]
        out[slot] = {"version": ver, "filename": path.name, "path": str(path)}
    return out


def local_latest_prompt2_versions(outputs_dir: Path, skus: list[str] | None = None) -> dict[str, int]:
    all_skus = skus or list_output_sku_dirs(outputs_dir)
    out: dict[str, int] = {}
    for sku in all_skus:
        versions = list_prompt_versions(sku_workspace_dir(outputs_dir, sku), "prompt2")
        if versions:
            out[sku] = versions[-1][0]
    return out


def drive_prompt_versions(filenames: list[str]) -> dict[str, list[int]]:
    grouped: dict[str, list[int]] = {"prompt1": [], "prompt2": []}
    for name in filenames:
        base = name.split("/")[-1]
        m = PROMPT_VERSION_FILE_RE.match(base)
        if not m:
            continue
        slot = "prompt1" if m.group(1) == "1" else "prompt2"
        grouped[slot].append(int(m.group(2)))
    for slot in grouped:
        grouped[slot].sort()
    return grouped


def drive_sync_needed(
    local: dict[str, dict[str, Any]],
    drive: dict[str, list[int]],
) -> tuple[bool, list[str]]:
    """True when Drive is missing local latest or still has older prompt versions."""
    issues: list[str] = []
    for slot in ("prompt1", "prompt2"):
        loc = local.get(slot)
        if not loc:
            continue
        want = int(loc["version"])
        remote = drive.get(slot) or []
        if want not in remote:
            issues.append(f"missing_{slot}_v{want}")
        stale = [v for v in remote if v != want]
        if stale:
            issues.append(f"stale_{slot}:{stale}")
    return bool(issues), issues


def shopify_stale_slots(
    local: dict[str, dict[str, Any]],
    review_rec: dict[str, Any],
) -> tuple[list[str], list[str]]:
    """
    Compare local latest versions to review_store approved versions (last known Shopify upload).
    Returns (stale_slots, issues).
    """
    stale: list[str] = []
    issues: list[str] = []
    for slot, approved_key in (("prompt1", "approved_prompt1_version"), ("prompt2", "approved_prompt2_version")):
        loc = local.get(slot)
        if not loc:
            continue
        local_ver = int(loc["version"])
        approved = review_rec.get(approved_key)
        approved_ver = int(approved) if approved is not None else None
        if approved_ver is None:
            issues.append(f"shopify_never_synced_{slot}")
            stale.append(slot)
        elif local_ver != approved_ver:
            issues.append(f"shopify_stale_{slot}:local_v{local_ver}_approved_v{approved_ver}")
            stale.append(slot)
    return stale, issues


def load_xlsx_thumb_paths(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.is_file():
        return out
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        sku_i = headers.index("SKU")
        path_i = headers.index("thumbnail image path")
    except ValueError:
        wb.close()
        return out
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_i] or "").strip()
        if sku:
            out[sku] = str(row[path_i] or "").strip()
    wb.close()
    return out


def expected_thumb_path(outputs_dir: Path, sku: str, *, prompt2_version: int | None) -> str:
    return thumbnail_relative_path(
        outputs_dir=outputs_dir,
        sku=sku,
        prompt2_version=prompt2_version,
    )


def audit_sku_prompt_versions(
    cfg: DriveReviewConfig,
    sync: DriveOutputsSync,
    sku: str,
    *,
    drive_folders: dict[str, str] | None = None,
    shop_prod: dict[str, Any] | None = None,
    review_store: ReviewStore | None = None,
) -> dict[str, Any]:
    review_store = review_store or ReviewStore(cfg.review_state_path)
    local = local_prompt_state(cfg.outputs_dir, sku)
    rec = review_store.get_record(sku)

    drive_meta: dict[str, Any] = {}
    drive_versions: dict[str, list[int]] = {"prompt1": [], "prompt2": []}
    if drive_folders is None:
        drive_folders = sync.list_sku_folders(refresh=False)
    if sku in drive_folders:
        drive_meta = sync.scan_sku_metadata(sku, folder_id=drive_folders[sku])
        drive_versions = drive_prompt_versions(drive_meta.get("filenames") or [])

    need_drive, drive_issues = drive_sync_needed(local, drive_versions)
    stale_shopify, shopify_issues = shopify_stale_slots(local, rec)

    shopify_images = 0
    if shop_prod:
        shopify_images = len(
            [
                m
                for m in shop_prod.get("media") or []
                if str(m.get("content_type") or "IMAGE").upper() not in {"VIDEO", "EXTERNAL_VIDEO"}
            ]
        )
    if shop_prod and local.get("prompt2") and shopify_images < 1:
        shopify_issues.append("shopify_missing_prompt2_slot")
        if "prompt2" not in stale_shopify:
            stale_shopify.append("prompt2")
    if shop_prod and local.get("prompt1") and shopify_images < 2:
        shopify_issues.append("shopify_missing_prompt1_slot")
        if "prompt1" not in stale_shopify:
            stale_shopify.append("prompt1")

    p2_ver = (local.get("prompt2") or {}).get("version")
    expected_xlsx = expected_thumb_path(cfg.outputs_dir, sku, prompt2_version=int(p2_ver) if p2_ver else None)

    return {
        "sku": sku,
        "local": local,
        "drive_versions": drive_versions,
        "drive_issues": drive_issues,
        "need_drive_sync": need_drive,
        "shopify_issues": shopify_issues,
        "shopify_stale_slots": stale_shopify,
        "shopify_product_id": str(rec.get("product_id") or (shop_prod or {}).get("id") or ""),
        "expected_xlsx_thumb_path": expected_xlsx,
    }


def sync_drive_prompts(
    sync: DriveOutputsSync,
    sku: str,
    *,
    local: dict[str, dict[str, Any]],
    dry_run: bool = False,
) -> dict[str, Any]:
    slots = [s for s in ("prompt1", "prompt2") if s in local]
    if dry_run:
        return {"sku": sku, "dry_run": True, "would_push_slots": slots}
    return sync.push_prompt_files(sku, slots=slots, prune_old=True)


def sync_shopify_prompts(
    client,
    cfg: DriveReviewConfig,
    sku: str,
    *,
    local: dict[str, dict[str, Any]],
    slots: list[str],
    shop_prod: dict[str, Any] | None,
    review_store: ReviewStore,
    dry_run: bool = False,
) -> dict[str, Any]:
    product_id = str(review_store.get_record(sku).get("product_id") or (shop_prod or {}).get("id") or "").strip()
    if not product_id:
        return {"sku": sku, "skipped": "no_product_id"}
    if dry_run:
        return {"sku": sku, "dry_run": True, "would_replace_slots": slots, "product_id": product_id}
    result = replace_prompt_images_on_product(
        client,
        product_id=product_id,
        sku=sku,
        slots=slots,
        shop_media=(shop_prod or {}).get("media"),
        review_store=review_store,
        cfg=cfg.base,
    )
    if not result.get("errors"):
        patch: dict[str, Any] = {}
        if local.get("prompt1"):
            patch["approved_prompt1_version"] = int(local["prompt1"]["version"])
        if local.get("prompt2"):
            patch["approved_prompt2_version"] = int(local["prompt2"]["version"])
        if patch:
            review_store.update(sku, **patch)
    return result


def audit_xlsx_alignment(
    cfg: DriveReviewConfig,
    *,
    skus: list[str],
    prompt2_versions: dict[str, int],
    local_enriched: Path,
    drive_enriched: Path | None = None,
) -> list[dict[str, Any]]:
    local_paths = load_xlsx_thumb_paths(local_enriched)
    drive_paths = load_xlsx_thumb_paths(drive_enriched) if drive_enriched else {}
    rows: list[dict[str, Any]] = []
    for sku in skus:
        expected = expected_thumb_path(
            cfg.outputs_dir,
            sku,
            prompt2_version=prompt2_versions.get(sku),
        )
        if not expected:
            continue
        local_actual = local_paths.get(sku, "")
        drive_actual = drive_paths.get(sku, "")
        issues: list[str] = []
        if local_actual != expected:
            issues.append(f"local_xlsx:{local_actual!r}!=expected:{expected!r}")
        if drive_enriched and drive_actual != expected:
            issues.append(f"drive_xlsx:{drive_actual!r}!=expected:{expected!r}")
        rows.append(
            {
                "sku": sku,
                "expected": expected,
                "local_xlsx": local_actual,
                "drive_xlsx": drive_actual,
                "ok": not issues,
                "issues": issues,
            }
        )
    return rows


def rebuild_enriched_from_local_latest(
    cfg: DriveReviewConfig,
    service,
    *,
    review_store: ReviewStore | None = None,
    prompt2_versions: dict[str, int] | None = None,
) -> Path:
    """Rebuild stock_enriched.xlsx using filesystem-latest prompt2 versions."""
    review_store = review_store or ReviewStore(cfg.review_state_path)
    if prompt2_versions is None:
        prompt2_versions = local_latest_prompt2_versions(cfg.outputs_dir)

    product_names = review_store.product_names()
    for sku, rec in review_store.all_records().items():
        title = str(rec.get("title") or "").strip()
        if title and sku not in product_names:
            product_names[sku] = title

    products_path = Path("products_export_1.xlsx")
    if not products_path.is_file():
        products_path = cfg.base.xlsx_path.parent / "products_export_1.xlsx"

    stock_path = resolve_stock_path(cfg, service)
    build_export(
        stock_path=stock_path,
        products_path=products_path,
        outputs_dir=cfg.outputs_dir,
        output_path=cfg.enriched_xlsx_path,
        stock_sheets=cfg.base.xlsx_sheets,
        product_names=product_names,
        images_dir=cfg.base.images_dir,
        prompt2_versions=prompt2_versions,
    )
    return cfg.enriched_xlsx_path


def run_prompt_version_sync(
    cfg: DriveReviewConfig,
    sync: DriveOutputsSync,
    service,
    *,
    shopify_client=None,
    products_by_sku: dict[str, dict] | None = None,
    review_store: ReviewStore | None = None,
    skus: list[str] | None = None,
    dry_run: bool = False,
    fix_drive: bool = True,
    fix_shopify: bool = True,
    fix_xlsx: bool = True,
    progress: Callable[[str, int, int], None] | None = None,
) -> dict[str, Any]:
    review_store = review_store or ReviewStore(cfg.review_state_path)
    products_by_sku = products_by_sku or {}

    if skus is None:
        skus = sorted(
            s
            for s in list_output_sku_dirs(cfg.outputs_dir)
            if local_prompt_state(cfg.outputs_dir, s)
        )

    drive_folders = sync.list_sku_folders(refresh=True)
    t0 = time.monotonic()

    audits: list[dict[str, Any]] = []
    drive_synced: list[str] = []
    drive_errors: list[dict[str, Any]] = []
    shopify_synced: list[str] = []
    shopify_errors: list[dict[str, Any]] = []

    total = len(skus)
    for i, sku in enumerate(skus, start=1):
        if progress:
            progress("Auditing SKU prompt versions", i, total)
        audit = audit_sku_prompt_versions(
            cfg,
            sync,
            sku,
            drive_folders=drive_folders,
            shop_prod=products_by_sku.get(sku),
            review_store=review_store,
        )
        audits.append(audit)

        if fix_drive and audit["need_drive_sync"]:
            try:
                sync_drive_prompts(sync, sku, local=audit["local"], dry_run=dry_run)
                if not dry_run:
                    drive_synced.append(sku)
            except Exception as e:
                drive_errors.append({"sku": sku, "error": str(e)})

        if fix_shopify and audit["shopify_stale_slots"] and shopify_client:
            try:
                result = sync_shopify_prompts(
                    shopify_client,
                    cfg,
                    sku,
                    local=audit["local"],
                    slots=audit["shopify_stale_slots"],
                    shop_prod=products_by_sku.get(sku),
                    review_store=review_store,
                    dry_run=dry_run,
                )
                if result.get("errors"):
                    shopify_errors.append({"sku": sku, "errors": result["errors"]})
                elif not dry_run and not result.get("skipped"):
                    shopify_synced.append(sku)
            except Exception as e:
                shopify_errors.append({"sku": sku, "error": str(e)})

    prompt2_versions = local_latest_prompt2_versions(cfg.outputs_dir, skus)
    xlsx_rows = audit_xlsx_alignment(
        cfg,
        skus=skus,
        prompt2_versions=prompt2_versions,
        local_enriched=cfg.enriched_xlsx_path,
        drive_enriched=resolve_stock_path(cfg, service) if service else None,
    )
    xlsx_mismatches = [r for r in xlsx_rows if not r["ok"]]

    xlsx_fix: dict[str, Any] = {}
    if fix_xlsx and xlsx_mismatches and service and not dry_run:
        rebuild_enriched_from_local_latest(cfg, service, review_store=review_store, prompt2_versions=prompt2_versions)
        xlsx_fix["local_rebuilt"] = str(cfg.enriched_xlsx_path)
        xlsx_fix["drive_replace"] = replace_drive_spreadsheet(cfg, service, local_path=cfg.enriched_xlsx_path)
        xlsx_rows = audit_xlsx_alignment(
            cfg,
            skus=skus,
            prompt2_versions=prompt2_versions,
            local_enriched=cfg.enriched_xlsx_path,
            drive_enriched=resolve_stock_path(cfg, service, force_refresh=True),
        )
        xlsx_mismatches = [r for r in xlsx_rows if not r["ok"]]

    need_drive = [a["sku"] for a in audits if a["need_drive_sync"]]
    need_shopify = [a["sku"] for a in audits if a["shopify_stale_slots"]]

    return {
        "generated_at": _now_utc(),
        "dry_run": dry_run,
        "elapsed_seconds": round(time.monotonic() - t0, 1),
        "sku_count": total,
        "summary": {
            "need_drive_sync": len(need_drive),
            "need_shopify_sync": len(need_shopify),
            "xlsx_mismatches": len(xlsx_mismatches),
            "drive_synced": len(drive_synced),
            "shopify_synced": len(shopify_synced),
            "drive_errors": len(drive_errors),
            "shopify_errors": len(shopify_errors),
        },
        "need_drive_sync": need_drive,
        "need_shopify_sync": need_shopify,
        "drive_synced": drive_synced,
        "shopify_synced": shopify_synced,
        "drive_errors": drive_errors,
        "shopify_errors": shopify_errors,
        "xlsx_rows": xlsx_rows,
        "xlsx_mismatches": xlsx_mismatches,
        "xlsx_fix": xlsx_fix,
        "audits": audits,
    }


def write_sync_report(report: dict[str, Any], outputs_dir: Path) -> Path:
    out_dir = outputs_dir / "rebuild_executor"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "prompt_version_sync_report.json"
    path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    return path
