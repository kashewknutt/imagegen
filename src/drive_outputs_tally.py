"""Compare local outputs/ with Drive SKU folders and cross-check Stock/review/Shopify."""
from __future__ import annotations

import json
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from src.drive_outputs_sync import DriveOutputsSync
from src.drive_review_config import DriveReviewConfig
from src.drive_review_log import get_logger
from src.drive_stock_sync import resolve_stock_path
from src.media_workspace import list_prompt_versions, sku_workspace_dir
from src.review_store import ReviewStore
from src.typo_sku_cleanup import KNOWN_TYPO_TO_REAL, load_stock_skus


def _now_utc() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def _local_flags(outputs_dir: Path, sku: str) -> dict[str, Any]:
    sku_dir = sku_workspace_dir(outputs_dir, sku)
    p1 = list_prompt_versions(sku_dir, "prompt1")
    p2 = list_prompt_versions(sku_dir, "prompt2")
    return {
        "has_prompt1": bool(p1),
        "has_prompt2": bool(p2),
        "prompt1_count": len(p1),
        "prompt2_count": len(p2),
    }


def _shopify_skus(shopify_client) -> set[str]:
    from src.shopify_product_dedup import fetch_all_shopify_products, primary_sku_from_product

    out: set[str] = set()
    for prod in fetch_all_shopify_products(shopify_client, active_only=False):
        sku = primary_sku_from_product(prod)
        if sku:
            out.add(sku)
    return out


def _enriched_skus(path: Path) -> set[str]:
    if not path.is_file():
        return set()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "SKU" not in headers:
            wb.close()
            return set()
        sku_i = headers.index("SKU")
        out: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku = str(row[sku_i] or "").strip()
            if sku:
                out.add(sku)
        wb.close()
        return out
    except Exception:
        return set()


def tally_drive_vs_local(
    cfg: DriveReviewConfig,
    sync: DriveOutputsSync,
    *,
    shopify_client=None,
    progress: Callable[[str, int, int], None] | None = None,
) -> dict[str, Any]:
    """Metadata-only Drive scan + local filesystem checks (no downloads)."""
    log = get_logger()
    t0 = time.monotonic()
    stock_path = resolve_stock_path(cfg, sync.service)
    stock_skus = load_stock_skus(stock_path)
    enriched_skus = _enriched_skus(cfg.enriched_xlsx_path)
    review_store = ReviewStore(cfg.review_state_path)

    local_skus = set(sync.list_local_sku_dirs())
    drive_map = sync.list_sku_folders(refresh=True)
    drive_skus = set(drive_map.keys())

    both = sorted(local_skus & drive_skus)
    local_only = sorted(local_skus - drive_skus)
    drive_only = sorted(drive_skus - local_skus)

    shopify_set: set[str] = set()
    shopify_error: str | None = None
    if shopify_client is not None:
        try:
            log.info("Loading Shopify SKU index for tally...")
            shopify_set = _shopify_skus(shopify_client)
        except Exception as e:
            shopify_error = str(e)
            log.warning("Shopify tally skipped: %s", e)

    def _sku_row(sku: str, *, on_drive: bool, on_local: bool) -> dict[str, Any]:
        local = _local_flags(cfg.outputs_dir, sku) if on_local else {}
        review_status = "no record"
        if on_local and sku in review_store.all_records():
            review_status = str(review_store.get_record(sku).get("review_status") or "pending_review")
        drive_meta: dict[str, Any] = {}
        if on_drive and sku in drive_map:
            drive_meta = sync.scan_sku_metadata(sku, folder_id=drive_map[sku])
        mismatch: list[str] = []
        if on_local and on_drive:
            if local.get("has_prompt1") and not drive_meta.get("has_prompt1"):
                mismatch.append("prompt1_local_only")
            if local.get("has_prompt2") and not drive_meta.get("has_prompt2"):
                mismatch.append("prompt2_local_only")
            if drive_meta.get("has_prompt1") and not local.get("has_prompt1"):
                mismatch.append("prompt1_drive_only")
            if drive_meta.get("has_prompt2") and not local.get("has_prompt2"):
                mismatch.append("prompt2_drive_only")
        return {
            "sku": sku,
            "in_stock": sku in stock_skus,
            "in_enriched": sku in enriched_skus,
            "in_review_state": sku in review_store.all_records(),
            "review_status": review_status,
            "on_shopify": sku in shopify_set if shopify_client else None,
            "on_local": on_local,
            "on_drive": on_drive,
            "local": local,
            "drive": {
                "has_prompt1": drive_meta.get("has_prompt1"),
                "has_prompt2": drive_meta.get("has_prompt2"),
                "raw_count": drive_meta.get("raw_count"),
                "video_count": drive_meta.get("video_count"),
                "file_count": len(drive_meta.get("filenames") or []),
            },
            "known_typo_of": KNOWN_TYPO_TO_REAL.get(sku),
            "mismatch": mismatch,
        }

    rows: list[dict[str, Any]] = []
    scan_skus = sorted(local_skus | drive_skus)
    total = len(scan_skus)
    for i, sku in enumerate(scan_skus, start=1):
        if progress:
            progress("Scanning SKU metadata", i, total)
        rows.append(_sku_row(sku, on_drive=sku in drive_skus, on_local=sku in local_skus))

    mismatched = [r for r in rows if r.get("mismatch")]
    not_in_stock_local = [r["sku"] for r in rows if r["on_local"] and not r["in_stock"]]
    local_missing_prompts = [
        r["sku"]
        for r in rows
        if r["on_local"] and (not (r.get("local") or {}).get("has_prompt1") or not (r.get("local") or {}).get("has_prompt2"))
    ]
    needs_push = [r["sku"] for r in mismatched if any(m.endswith("_local_only") for m in r["mismatch"])]

    review_counts: dict[str, int] = {}
    for sku in local_skus:
        if sku not in review_store.all_records():
            review_counts["no_record"] = review_counts.get("no_record", 0) + 1
            continue
        status = str(review_store.get_record(sku).get("review_status") or "pending_review")
        review_counts[status] = review_counts.get(status, 0) + 1

    elapsed = round(time.monotonic() - t0, 1)
    summary = {
        "local_sku_folders": len(local_skus),
        "drive_sku_folders": len(drive_skus),
        "in_both": len(both),
        "local_only": len(local_only),
        "drive_only": len(drive_only),
        "stock_xlsx_skus": len(stock_skus),
        "enriched_xlsx_skus": len(enriched_skus),
        "local_not_in_stock": len(not_in_stock_local),
        "local_missing_prompts": len(local_missing_prompts),
        "content_mismatches": len(mismatched),
        "needs_push_to_drive": len(needs_push),
        "shopify_skus": len(shopify_set) if shopify_client else None,
        "review_status_counts": review_counts,
        "elapsed_seconds": elapsed,
    }
    if shopify_error:
        summary["shopify_error"] = shopify_error

    log.info(
        "Tally complete in %.1fs: local=%d drive=%d both=%d local_only=%d drive_only=%d mismatches=%d",
        elapsed,
        len(local_skus),
        len(drive_skus),
        len(both),
        len(local_only),
        len(drive_only),
        len(mismatched),
    )

    return {
        "generated_at": _now_utc(),
        "stock_path": str(stock_path),
        "local_outputs_dir": str(cfg.outputs_dir),
        "drive_outputs_folder_id": cfg.drive_outputs_folder_id,
        "stock_spreadsheet_id": cfg.stock_spreadsheet_id,
        "summary": summary,
        "local_only": local_only,
        "drive_only": drive_only,
        "needs_push_to_drive": needs_push,
        "not_in_stock_local": not_in_stock_local,
        "local_missing_prompts": local_missing_prompts,
        "content_mismatches": mismatched,
        "rows": rows,
    }


def write_tally_report(tally: dict[str, Any], outputs_dir: Path) -> tuple[Path, Path]:
    outputs_dir.mkdir(parents=True, exist_ok=True)
    json_path = outputs_dir / "drive_tally_report.json"
    md_path = outputs_dir / "drive_tally_report.md"
    json_path.write_text(json.dumps(tally, indent=2, sort_keys=True), encoding="utf-8")

    s = tally.get("summary") or {}
    lines = [
        "# Drive vs local outputs tally",
        "",
        f"Generated: {tally.get('generated_at', '')}",
        "",
        "## Summary",
        "",
        f"- Local SKU folders: **{s.get('local_sku_folders', 0)}**",
        f"- Drive SKU folders: **{s.get('drive_sku_folders', 0)}**",
        f"- In both: **{s.get('in_both', 0)}**",
        f"- Local only (not on Drive): **{s.get('local_only', 0)}**",
        f"- Drive only (not local): **{s.get('drive_only', 0)}**",
        f"- Stock.xlsx SKUs: **{s.get('stock_xlsx_skus', 0)}**",
        f"- Local folders not in Stock: **{s.get('local_not_in_stock', 0)}**",
        f"- Local missing prompt1 or prompt2: **{s.get('local_missing_prompts', 0)}**",
        f"- Content mismatches (local vs Drive metadata): **{s.get('content_mismatches', 0)}**",
        f"- Needs push to Drive: **{s.get('needs_push_to_drive', 0)}**",
        "",
    ]
    if s.get("shopify_skus") is not None:
        lines.append(f"- Shopify SKUs: **{s.get('shopify_skus', 0)}**")
    if s.get("shopify_error"):
        lines.append(f"- Shopify error: `{s['shopify_error']}`")
    lines.extend(["", f"Elapsed: {s.get('elapsed_seconds', '?')}s", ""])

    def _list_section(title: str, items: list[str], limit: int = 40) -> None:
        lines.append(f"## {title} ({len(items)})")
        lines.append("")
        if not items:
            lines.append("(none)")
        else:
            for sku in items[:limit]:
                lines.append(f"- `{sku}`")
            if len(items) > limit:
                lines.append(f"- … and {len(items) - limit} more")
        lines.append("")

    _list_section("Local only", tally.get("local_only") or [])
    _list_section("Drive only", tally.get("drive_only") or [])
    _list_section("Not in Stock.xlsx (local)", tally.get("not_in_stock_local") or [])
    _list_section("Needs push to Drive", tally.get("needs_push_to_drive") or [])

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path
