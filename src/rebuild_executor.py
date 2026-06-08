"""Phased executor for Shopify wipe, local prune, validation, and approved reupload."""
from __future__ import annotations

import json
import logging
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from build_stock_export import build_export
from dedupe_duplicate_skus import audit_stock
from src.media_workspace import (
    index_sku_media,
    prune_all_prompt_versions,
    sku_workspace_dir,
    thumbnail_relative_path,
)
from src.review_store import ReviewStore
from src.shopify_media_sync import images_for_sku, media_paths_for_sku, sync_product_media
from src.shopify_product_dedup import (
    fetch_all_shopify_products,
    group_shopify_products_by_sku,
    primary_sku_from_product,
)
from src.title_store import TitleStore
from src.upload_store import UploadStore
from src.xlsx_ingest import index_by_sku, iter_rows

log = logging.getLogger(__name__)

APPROVED_REVIEW_STATUSES = {"approved", "uploaded"}
REPORT_SUBDIR = "rebuild_executor"


def _now_utc() -> str:
    return datetime.now(tz=timezone.utc).isoformat()


def _norm_title(title: str) -> str:
    return " ".join((title or "").strip().lower().split())


def _report_dir(outputs_dir: Path) -> Path:
    path = outputs_dir / REPORT_SUBDIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def save_phase_report(outputs_dir: Path, phase: str, data: dict[str, Any]) -> Path:
    path = _report_dir(outputs_dir) / f"{phase}.json"
    payload = {"phase": phase, "timestamp_utc": _now_utc(), **data}
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return path


@dataclass
class EligibilityResult:
    eligible: list[str] = field(default_factory=list)
    blocked_local: dict[str, list[str]] = field(default_factory=dict)
    blocked_stock: list[str] = field(default_factory=list)
    not_approved: list[str] = field(default_factory=list)


def verify_shopify_connection(client) -> dict[str, Any]:
    page = client.list_products(first=1, after=None, query=None)
    products = page.get("products") or []
    return {"ok": True, "sample_count": len(products)}


def verify_product_delete_supported(client) -> None:
    """Fail loudly if productDelete mutation is unavailable."""
    mutation = """
    mutation ProductDelete($input: ProductDeleteInput!) {
      productDelete(input: $input) {
        deletedProductId
        userErrors { field message }
      }
    }
    """
    data = client.graphql(mutation, {"input": {"id": "gid://shopify/Product/0"}})
    if "productDelete" not in (data or {}):
        raise RuntimeError(
            "Shopify GraphQL schema does not expose productDelete — hard delete is required for wipe."
        )


def delete_all_shopify_products(
    client,
    *,
    outputs_dir: Path,
    dry_run: bool = False,
    sleep_seconds: float = 0.35,
) -> dict[str, Any]:
    verify_product_delete_supported(client)
    products = fetch_all_shopify_products(client, active_only=False)
    deleted: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []

    log.info("Found %d Shopify product(s) to delete (all statuses)", len(products))
    for i, prod in enumerate(products, start=1):
        pid = str(prod.get("id") or "")
        handle = str(prod.get("handle") or "")
        title = str(prod.get("title") or "")
        sku = primary_sku_from_product(prod)
        log.info("[%d/%d] delete %s (%s) sku=%s", i, len(products), handle or pid, title[:60], sku)
        if dry_run:
            deleted.append({"product_id": pid, "handle": handle, "sku": sku, "title": title})
            continue
        try:
            client.product_delete(product_id=pid)
            deleted.append({"product_id": pid, "handle": handle, "sku": sku, "title": title})
        except Exception as e:
            errors.append({"product_id": pid, "handle": handle, "sku": sku, "error": str(e)})
            log.error("Failed to delete %s: %s", pid, e)
        time.sleep(sleep_seconds)

    report = {
        "dry_run": dry_run,
        "total_found": len(products),
        "deleted_count": len(deleted),
        "error_count": len(errors),
        "deleted": deleted,
        "errors": errors,
    }
    save_phase_report(outputs_dir, "delete_all_products", report)
    if errors and not dry_run:
        raise RuntimeError(f"Shopify delete phase failed for {len(errors)} product(s)")
    return report


def reset_local_shopify_identity(outputs_dir: Path) -> dict[str, Any]:
    review_store = ReviewStore(outputs_dir / "review_state.json")
    title_store = TitleStore(outputs_dir / "title_gen_state.json")
    upload_store = UploadStore(outputs_dir / "upload_state.json")

    review_reset = 0
    for sku, rec in review_store.all_records().items():
        needs = (
            str(rec.get("product_id") or "").strip()
            or str(rec.get("handle") or "").strip()
            or list(rec.get("shopify_media_ids") or [])
            or str(rec.get("upload_status") or "pending") not in {"", "pending"}
            or str(rec.get("last_error") or "").strip()
        )
        if not needs:
            continue
        status = str(rec.get("review_status") or "pending_review")
        upload_status = "pending"
        if status not in APPROVED_REVIEW_STATUSES:
            upload_status = str(rec.get("upload_status") or "pending")
        patch: dict[str, Any] = {
            "product_id": "",
            "handle": "",
            "shopify_media_ids": [],
            "upload_status": upload_status,
            "last_error": "",
        }
        if status == "uploaded":
            patch["review_status"] = "approved"
        review_store.update(sku, **patch)
        review_reset += 1

    title_reset = 0
    for key, rec in title_store.all_records().items():
        if str(rec.get("product_id") or "").strip() or str(rec.get("handle") or "").strip():
            title_store.update(key, product_id="", handle="")
            title_reset += 1

    upload_reset = 0
    for sku, rec in upload_store.all_records().items():
        if (
            str(rec.get("product_id") or "").strip()
            or str(rec.get("handle") or "").strip()
            or str(rec.get("status") or "pending") not in {"pending", ""}
            or str(rec.get("last_error") or "").strip()
        ):
            upload_store.update(sku, product_id="", handle="", status="pending", last_error="")
            upload_reset += 1

    report = {
        "review_records_reset": review_reset,
        "title_records_reset": title_reset,
        "upload_records_reset": upload_reset,
    }
    save_phase_report(outputs_dir, "reset_local_state", report)
    return report


def run_prompt_prune(
    outputs_dir: Path,
    *,
    review_store: ReviewStore | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    results = prune_all_prompt_versions(outputs_dir, review_store=review_store, dry_run=dry_run)
    total_deleted = sum(len(r.get("deleted") or []) for r in results)
    skus_pruned = sum(1 for r in results if r.get("deleted"))
    report = {
        "dry_run": dry_run,
        "skus_scanned": len(results),
        "skus_pruned": skus_pruned,
        "files_deleted": total_deleted,
        "details": results,
    }
    save_phase_report(outputs_dir, "prune_prompt_versions", report)
    return report


def _local_asset_issues(cfg, sku: str, review_store: ReviewStore) -> list[str]:
    issues: list[str] = []
    ws_dir = sku_workspace_dir(cfg.outputs_dir, sku)
    if not ws_dir.is_dir():
        issues.append("missing workspace folder")
        return issues
    idx = index_sku_media(outputs_dir=cfg.outputs_dir, sku=sku)
    paths = media_paths_for_sku(cfg, sku, review_store=review_store)
    if not idx.raw_images:
        issues.append("missing raw images")
    if not paths.get("prompt1"):
        issues.append("missing prompt1 image")
    if not paths.get("prompt2"):
        issues.append("missing prompt2 image")
    return issues


def build_eligibility(
    cfg,
    review_store: ReviewStore,
    stock_path: Path,
) -> EligibilityResult:
    stock_rows = index_by_sku(iter_rows(stock_path, ["Total"]), sku_column="SKU")
    result = EligibilityResult()

    for sku, rec in sorted(review_store.all_records().items()):
        status = str(rec.get("review_status") or "pending_review")
        if status not in APPROVED_REVIEW_STATUSES:
            result.not_approved.append(sku)
            continue
        if sku not in stock_rows:
            result.blocked_stock.append(sku)
            continue
        issues = _local_asset_issues(cfg, sku, review_store)
        if issues:
            result.blocked_local[sku] = issues
            continue
        result.eligible.append(sku)

    report = {
        "eligible_count": len(result.eligible),
        "eligible": result.eligible,
        "blocked_local": result.blocked_local,
        "blocked_stock": result.blocked_stock,
        "not_approved": result.not_approved,
    }
    save_phase_report(cfg.outputs_dir, "eligibility", report)
    return result


def run_uniqueness_checks(
    cfg,
    review_store: ReviewStore,
    stock_path: Path,
    eligible_skus: list[str],
) -> dict[str, Any]:
    stock_dupes = audit_stock(stock_path)
    eligible_set = set(eligible_skus)
    stock_dupes_eligible = [(sku, n) for sku, n in stock_dupes if sku in eligible_set]

    title_map: dict[str, list[str]] = defaultdict(list)
    missing_titles: list[str] = []
    for sku in eligible_skus:
        title = str(review_store.get_record(sku).get("title") or "").strip()
        if not title:
            missing_titles.append(sku)
            continue
        title_map[_norm_title(title)].append(sku)
    title_dupes = {title: skus for title, skus in title_map.items() if len(skus) > 1}

    workspace_counts = Counter(
        p.name
        for p in cfg.outputs_dir.iterdir()
        if p.is_dir() and not p.name.startswith(".") and p.name in eligible_set
    )
    workspace_dupes = [(sku, n) for sku, n in workspace_counts.items() if n > 1]

    report = {
        "eligible_count": len(eligible_skus),
        "stock_duplicate_skus": stock_dupes,
        "stock_duplicate_skus_in_eligible": stock_dupes_eligible,
        "title_duplicate_groups": title_dupes,
        "missing_titles": missing_titles,
        "workspace_duplicate_skus": workspace_dupes,
        "passed": not stock_dupes_eligible and not title_dupes and not missing_titles,
    }
    save_phase_report(cfg.outputs_dir, "uniqueness_checks", report)
    if stock_dupes_eligible:
        raise RuntimeError(f"Duplicate SKUs in Stock.xlsx for eligible set: {stock_dupes_eligible}")
    if title_dupes:
        raise RuntimeError(f"Duplicate approved titles: {title_dupes}")
    if missing_titles:
        raise RuntimeError(f"Approved eligible SKU(s) missing title: {missing_titles[:20]}")
    return report


def _load_xlsx_maps(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    from openpyxl import load_workbook

    names: dict[str, str] = {}
    paths: dict[str, str] = {}
    if not path.is_file():
        return names, paths
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        sku_i = headers.index("SKU")
        name_i = headers.index("productName")
        path_i = headers.index("thumbnail image path")
    except ValueError:
        wb.close()
        return names, paths
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_i] or "").strip()
        if not sku:
            continue
        names[sku] = str(row[name_i] or "").strip()
        paths[sku] = str(row[path_i] or "").strip()
    wb.close()
    return names, paths


def rebuild_stock_enriched(
    cfg,
    review_store: ReviewStore,
    *,
    stock_path: Path,
    products_path: Path,
    output_path: Path,
) -> int:
    prompt2_versions: dict[str, int | None] = {}
    for sku, rec in review_store.all_records().items():
        p2 = rec.get("approved_prompt2_version")
        if p2 is not None:
            prompt2_versions[sku] = int(p2)
    row_count = build_export(
        stock_path=stock_path,
        products_path=products_path,
        outputs_dir=cfg.outputs_dir,
        output_path=output_path,
        stock_sheets=cfg.xlsx_sheets,
        product_names=review_store.product_names(),
        images_dir=cfg.images_dir,
        prompt2_versions=prompt2_versions,
    )
    save_phase_report(
        cfg.outputs_dir,
        "rebuild_xlsx",
        {"output_path": str(output_path), "row_count": row_count},
    )
    return row_count


def validate_xlsx_for_eligible(
    cfg,
    review_store: ReviewStore,
    xlsx_path: Path,
    eligible_skus: list[str],
) -> dict[str, Any]:
    names, paths = _load_xlsx_maps(xlsx_path)
    errors: dict[str, list[str]] = {}

    for sku in eligible_skus:
        sku_errors: list[str] = []
        rec = review_store.get_record(sku)
        approved_title = str(rec.get("title") or "").strip()
        p2v = rec.get("approved_prompt2_version")
        expected_thumb = thumbnail_relative_path(
            outputs_dir=cfg.outputs_dir,
            sku=sku,
            prompt2_version=int(p2v) if p2v is not None else None,
        )
        xlsx_name = names.get(sku, "")
        xlsx_path_val = paths.get(sku, "")

        if not xlsx_name:
            sku_errors.append("missing productName in XLSX")
        elif approved_title and xlsx_name != approved_title:
            sku_errors.append(f"productName mismatch: '{xlsx_name}' vs '{approved_title}'")
        if not xlsx_path_val:
            sku_errors.append("missing thumbnail image path in XLSX")
        elif expected_thumb and xlsx_path_val != expected_thumb:
            sku_errors.append(f"thumbnail path mismatch: '{xlsx_path_val}' vs '{expected_thumb}'")

        p2_ver = int(p2v) if p2v is not None else None
        if p2_ver is not None:
            ws_dir = sku_workspace_dir(cfg.outputs_dir, sku)
            from src.media_workspace import list_prompt_versions

            kept_versions = {v for v, _ in list_prompt_versions(ws_dir, "prompt2")}
            if kept_versions and p2_ver not in kept_versions:
                sku_errors.append(
                    f"approved prompt2 v{p2_ver} not present after prune (have {sorted(kept_versions)})"
                )

        if sku_errors:
            errors[sku] = sku_errors

    report = {
        "xlsx_path": str(xlsx_path),
        "eligible_count": len(eligible_skus),
        "error_count": len(errors),
        "errors": errors,
        "passed": not errors,
    }
    save_phase_report(cfg.outputs_dir, "validate_xlsx", report)
    if errors:
        sample = {k: v for k, v in list(errors.items())[:5]}
        raise RuntimeError(f"XLSX validation failed for {len(errors)} SKU(s). Sample: {sample}")
    return report


def _parse_tags(tags_str: str) -> list[str]:
    return sorted({t.strip() for t in tags_str.replace(";", ",").split(",") if t.strip()})


def _gid_to_int(gid: str) -> int | None:
    try:
        s = str(gid or "").strip()
        return int(s.rsplit("/", 1)[-1]) if s else None
    except Exception:
        return None


def _parse_float(value: object) -> float | None:
    try:
        s = str(value).strip().replace(",", "")
        return float(s) if s else None
    except Exception:
        return None


def _enriched_row_index(enriched_path: Path) -> dict[str, dict]:
    from openpyxl import load_workbook

    if not enriched_path.is_file():
        return {}
    wb = load_workbook(enriched_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        sku = str(vals.get("SKU") or "").strip()
        if sku and sku not in out:
            out[sku] = vals
    wb.close()
    return out


def _stock_row(stock_path: Path, sku: str, *, enriched_path: Path | None = None) -> dict:
    rows = index_by_sku(iter_rows(stock_path, ["Total"]), sku_column="SKU")
    row = rows.get(sku)
    if row:
        return dict(getattr(row, "values", {}) or {})
    if enriched_path is not None:
        return dict(_enriched_row_index(enriched_path).get(sku) or {})
    return {}


def _is_transient_network_error(error: str) -> bool:
    err = (error or "").lower()
    markers = (
        "nameresolutionerror",
        "connecttimeouterror",
        "connection reset",
        "connection aborted",
        "max retries exceeded",
        "timed out",
        "no route to host",
        "failed to resolve",
    )
    return any(m in err for m in markers)


def list_remaining_upload_skus(
    cfg,
    review_store: ReviewStore,
    *,
    stock_path: Path | None = None,
) -> list[dict[str, str]]:
    """Eligible approved SKUs not yet marked uploaded (the main backlog)."""
    eligibility = build_eligibility(cfg, review_store, stock_path or cfg.xlsx_path)
    out: list[dict[str, str]] = []
    for sku in eligibility.eligible:
        rec = review_store.get_record(sku)
        if str(rec.get("review_status") or "") == "uploaded":
            continue
        out.append(
            {
                "sku": sku,
                "title": str(rec.get("title") or ""),
                "review_status": str(rec.get("review_status") or ""),
                "product_id": str(rec.get("product_id") or ""),
            }
        )
    return out


def upload_remaining_approved_products(
    cfg,
    client,
    *,
    skus: list[str] | None = None,
    dry_run: bool = False,
    sleep_seconds: float = 1.0,
    limit: int = 0,
    stop_after_network_errors: int = 3,
) -> dict[str, Any]:
    """Upload eligible SKUs that are approved/failed but not yet uploaded."""
    review_store = ReviewStore(cfg.outputs_dir / "review_state.json")
    title_store = TitleStore(cfg.outputs_dir / "title_gen_state.json")

    if skus:
        targets = [s.strip() for s in skus if s.strip()]
    else:
        targets = [item["sku"] for item in list_remaining_upload_skus(cfg, review_store)]

    if limit > 0:
        targets = targets[:limit]

    if not targets:
        log.info("No remaining SKUs to upload")
        return {"dry_run": dry_run, "targets": [], "success_count": 0, "failed_count": 0}

    log.info("Uploading %d remaining SKU(s) (approved backlog, not yet on Shopify)", len(targets))
    report = recreate_approved_products(
        cfg,
        client,
        review_store,
        title_store,
        cfg.xlsx_path,
        targets,
        dry_run=dry_run,
        sleep_seconds=sleep_seconds,
        stop_after_network_errors=stop_after_network_errors,
    )
    report["phase"] = "upload_remaining_approved"
    save_phase_report(cfg.outputs_dir, "upload_remaining_approved", report)
    return report


def list_failed_skus(review_store: ReviewStore) -> list[dict[str, str]]:
    """Return failed SKU records sorted by SKU."""
    out: list[dict[str, str]] = []
    for sku, rec in sorted(review_store.all_records().items()):
        if str(rec.get("review_status") or "") != "failed":
            continue
        out.append(
            {
                "sku": sku,
                "title": str(rec.get("title") or ""),
                "product_id": str(rec.get("product_id") or ""),
                "last_error": str(rec.get("last_error") or "")[:300],
            }
        )
    return out


def retry_failed_products(
    cfg,
    client,
    *,
    skus: list[str] | None = None,
    dry_run: bool = False,
    sleep_seconds: float = 1.0,
    limit: int = 0,
) -> dict[str, Any]:
    """Retry Shopify create/upload for SKUs marked failed in review_state."""
    review_store = ReviewStore(cfg.outputs_dir / "review_state.json")
    title_store = TitleStore(cfg.outputs_dir / "title_gen_state.json")

    if skus:
        targets = [s.strip() for s in skus if s.strip()]
    else:
        targets = [item["sku"] for item in list_failed_skus(review_store)]

    if limit > 0:
        targets = targets[:limit]

    if not targets:
        log.info("No failed SKUs to retry")
        return {"dry_run": dry_run, "targets": [], "success_count": 0, "failed_count": 0}

    log.info("Retrying %d failed SKU(s)", len(targets))
    for sku in targets:
        rec = review_store.get_record(sku)
        if str(rec.get("review_status") or "") == "failed":
            review_store.update(
                sku,
                review_status="approved",
                upload_status="pending",
                last_error="",
            )

    report = recreate_approved_products(
        cfg,
        client,
        review_store,
        title_store,
        cfg.xlsx_path,
        targets,
        dry_run=dry_run,
        sleep_seconds=sleep_seconds,
        stop_after_network_errors=3,
    )
    report["phase"] = "retry_failed_products"
    save_phase_report(cfg.outputs_dir, "retry_failed_products", report)
    return report


def recover_orphan_product_ids(
    client,
    review_store: ReviewStore,
    eligible_skus: list[str],
) -> int:
    """Link failed/interrupted SKUs to products already created on Shopify."""
    products = fetch_all_shopify_products(client, active_only=True)
    by_sku: dict[str, dict[str, Any]] = {}
    for prod in products:
        sku = primary_sku_from_product(prod)
        if sku and sku not in by_sku:
            by_sku[sku] = prod

    recovered = 0
    for sku in eligible_skus:
        rec = review_store.get_record(sku)
        if str(rec.get("product_id") or "").strip():
            continue
        prod = by_sku.get(sku)
        if not prod:
            continue
        review_store.update(
            sku,
            product_id=str(prod.get("id") or ""),
            handle=str(prod.get("handle") or ""),
            last_error="",
        )
        recovered += 1
        log.info("[%s] Recovered product_id from Shopify: %s", sku, prod.get("handle") or prod.get("id"))
    return recovered


def recreate_approved_products(
    cfg,
    client,
    review_store: ReviewStore,
    title_store: TitleStore,
    stock_path: Path,
    eligible_skus: list[str],
    *,
    dry_run: bool = False,
    sleep_seconds: float = 0.5,
    stop_after_network_errors: int = 0,
) -> dict[str, Any]:
    from upload_missing_shopify_products import _price_fields, _shopify_product_type

    ok: list[str] = []
    failed: dict[str, str] = {}
    created: list[dict[str, str]] = []
    stopped_early = False
    consecutive_network_errors = 0

    if not dry_run:
        recovered = recover_orphan_product_ids(client, review_store, eligible_skus)
        if recovered:
            log.info("Recovered %d orphan product_id(s) from Shopify for resume", recovered)

    for i, sku in enumerate(eligible_skus, start=1):
        rec = review_store.get_record(sku)
        title = str(rec.get("title") or "").strip()
        if not title:
            failed[sku] = "missing title"
            continue
        existing_pid = str(rec.get("product_id") or "").strip()
        if not dry_run and existing_pid and str(rec.get("review_status") or "") == "uploaded":
            log.info("[%d/%d] %s — already uploaded, skipping", i, len(eligible_skus), sku)
            ok.append(sku)
            continue

        images = images_for_sku(cfg, sku, review_store=review_store)
        if not images:
            failed[sku] = "no local images"
            review_store.mark_failed(sku, "no local images")
            continue

        row = _stock_row(stock_path, sku, enriched_path=cfg.outputs_dir / "stock_enriched.xlsx")
        category = str(rec.get("category") or row.get("category") or "").strip()
        product_type = str(rec.get("product_type") or "").strip() or _shopify_product_type(category, title=title)
        description = str(rec.get("description") or "").strip() or f"{title}."
        tags = _parse_tags(str(rec.get("tags") or ""))
        subcategory = str(row.get("subCategory") or "").strip()
        if subcategory and subcategory not in tags:
            tags.append(subcategory.title())
        price_sell, price_cost, weight_g, qty = _price_fields(row)

        log.info("[%d/%d] Creating %s: '%s'", i, len(eligible_skus), sku, title)
        if dry_run:
            created.append({"sku": sku, "title": title, "dry_run": "true"})
            ok.append(sku)
            continue

        try:
            if existing_pid:
                product_id = existing_pid
                handle = str(rec.get("handle") or "")
                log.info("[%d/%d] %s — resuming media sync for %s", i, len(eligible_skus), sku, product_id)
            else:
                prod = client.product_create(
                    title=title,
                    description_html=description,
                    vendor="ZOCI",
                    product_type=product_type,
                    tags=tags,
                )
                product_id = str(prod.get("id") or "")
                handle = str(prod.get("handle") or "")
                variant_id = _gid_to_int(prod.get("variant_id") or "")
                inventory_item_id = _gid_to_int(prod.get("inventory_item_id") or "")

                if variant_id:
                    client.rest_variant_update(variant_id=variant_id, sku=sku, price=price_sell or None)
                    if weight_g is not None:
                        client.rest_variant_weight(variant_id=variant_id, weight_kg=float(weight_g) / 1000.0)
                if inventory_item_id:
                    cost_f = _parse_float(price_cost)
                    if cost_f is not None:
                        client.rest_inventory_item_cost(inventory_item_id=inventory_item_id, cost=cost_f)
                    if qty:
                        locs = client.rest_locations()
                        if locs:
                            location_id = int((locs[0] or {}).get("id") or 0)
                            if location_id:
                                client.rest_inventory_set(
                                    location_id=location_id,
                                    inventory_item_id=inventory_item_id,
                                    available=int(qty),
                                )

                # Persist identity before media sync so interrupted runs can resume.
                review_store.update(sku, product_id=product_id, handle=handle, last_error="")

            paths = media_paths_for_sku(cfg, sku, review_store=review_store)
            existing_media = list(rec.get("shopify_media_ids") or [])
            media_result = sync_product_media(
                client,
                product_id=product_id,
                sku=sku,
                images=images,
                videos=paths.get("videos") or [],
                replace_existing=bool(existing_media),
                existing_media_ids=existing_media or None,
            )

            review_store.update(
                sku,
                product_id=product_id,
                handle=handle,
                shopify_media_ids=media_result.get("media_ids") or [],
            )
            review_store.mark_uploaded(sku, shopify_media_ids=media_result.get("media_ids") or [])
            title_store.update(
                sku,
                sku=sku,
                product_id=product_id,
                handle=handle,
                new_title=title,
                status="uploaded",
            )
            ok.append(sku)
            created.append({"sku": sku, "title": title, "product_id": product_id, "handle": handle})
            log.info("[%s] Created -> %s (%d images)", sku, handle or product_id, media_result.get("image_count", 0))
        except Exception as e:
            err = str(e)
            failed[sku] = err
            review_store.mark_failed(sku, err)
            log.error("[%s] Create failed: %s", sku, e)
            if _is_transient_network_error(err):
                consecutive_network_errors += 1
                if (
                    stop_after_network_errors > 0
                    and consecutive_network_errors >= stop_after_network_errors
                ):
                    log.error(
                        "Stopping after %d consecutive network error(s) — fix connection and re-run",
                        consecutive_network_errors,
                    )
                    stopped_early = True
                    break
            else:
                consecutive_network_errors = 0
        else:
            consecutive_network_errors = 0
        time.sleep(sleep_seconds)

    report = {
        "dry_run": dry_run,
        "eligible_count": len(eligible_skus),
        "success_count": len(ok),
        "failed_count": len(failed),
        "stopped_early": stopped_early,
        "success": ok,
        "failed": failed,
        "created": created,
    }
    save_phase_report(cfg.outputs_dir, "recreate_products", report)
    return report


def run_final_audit(
    cfg,
    review_store: ReviewStore,
    eligible_skus: list[str],
    *,
    client=None,
) -> dict[str, Any]:
    from audit_workspace import audit_sku

    xlsx_path = cfg.outputs_dir / "stock_enriched.xlsx"
    xlsx_names, xlsx_paths = _load_xlsx_maps(xlsx_path)

    shopify_by_sku: dict[str, dict] = {}
    shopify_errors: list[str] = []
    if client is not None:
        try:
            products = fetch_all_shopify_products(client, active_only=True)
            groups = group_shopify_products_by_sku(products)
            for sku, prods in groups.items():
                if prods:
                    shopify_by_sku[sku] = prods[0]
            dupes = {sku: len(prods) for sku, prods in groups.items() if len(prods) > 1}
            if dupes:
                shopify_errors.append(f"duplicate active Shopify SKUs: {dupes}")
        except Exception as e:
            shopify_errors.append(f"Shopify fetch failed: {e}")

    sku_errors: dict[str, list[str]] = {}
    sku_warnings: dict[str, list[str]] = {}
    for sku in eligible_skus:
        errors, warnings = audit_sku(
            cfg=cfg,
            sku=sku,
            shopify_product=shopify_by_sku.get(sku),
            review_store=review_store,
            xlsx_names=xlsx_names,
            xlsx_paths=xlsx_paths,
            strict_shopify=True,
        )
        if errors:
            sku_errors[sku] = errors
        if warnings:
            sku_warnings[sku] = warnings

    active_shopify_skus = set(shopify_by_sku)
    eligible_set = set(eligible_skus)
    missing_on_shopify = sorted(eligible_set - active_shopify_skus)
    extra_on_shopify = sorted(active_shopify_skus - eligible_set)

    report = {
        "eligible_count": len(eligible_skus),
        "shopify_active_count": len(shopify_by_sku),
        "shopify_sku_count_matches_eligible": len(shopify_by_sku) == len(eligible_skus),
        "missing_on_shopify": missing_on_shopify,
        "extra_on_shopify": extra_on_shopify[:50],
        "duplicate_shopify_skus": shopify_errors,
        "sku_error_count": len(sku_errors),
        "sku_errors": sku_errors,
        "sku_warnings": sku_warnings,
        "passed": (
            not sku_errors
            and not shopify_errors
            and not missing_on_shopify
            and len(shopify_by_sku) == len(eligible_skus)
        ),
    }
    save_phase_report(cfg.outputs_dir, "final_audit", report)
    return report


@dataclass
class ExecutorSummary:
    phases: dict[str, Any] = field(default_factory=dict)
    success: bool = False
    stopped_at: str = ""
    error: str = ""


def run_full_executor(
    cfg,
    client,
    *,
    confirm_delete_all: bool = False,
    dry_run: bool = False,
    skip_delete: bool = False,
    skip_reset: bool = False,
    skip_prune: bool = False,
    skip_upload: bool = False,
    products_path: Path | None = None,
) -> ExecutorSummary:
    summary = ExecutorSummary()
    stock_path = cfg.xlsx_path
    enriched_path = cfg.outputs_dir / "stock_enriched.xlsx"
    products_path = products_path or Path("products_export_1.xlsx")
    review_store = ReviewStore(cfg.outputs_dir / "review_state.json")
    title_store = TitleStore(cfg.outputs_dir / "title_gen_state.json")

    try:
        phase = "verify_client"
        summary.phases[phase] = verify_shopify_connection(client)
        log.info("Shopify connection OK")

        if not skip_delete:
            if not dry_run and not confirm_delete_all:
                raise RuntimeError("Refusing wipe: pass --confirm-delete-all to delete every Shopify product")
            phase = "delete_all_products"
            summary.phases[phase] = delete_all_shopify_products(
                client, outputs_dir=cfg.outputs_dir, dry_run=dry_run
            )
        else:
            log.info("Skipping Shopify delete phase")

        phase = "reset_local_state"
        if dry_run or skip_reset:
            summary.phases[phase] = {"skipped": True, "dry_run": dry_run, "skip_reset": skip_reset}
            if skip_reset:
                log.info("Skipping local state reset")
        else:
            summary.phases[phase] = reset_local_shopify_identity(cfg.outputs_dir)

        phase = "prune_prompt_versions"
        if skip_prune:
            summary.phases[phase] = {"skipped": True}
            log.info("Skipping prompt version prune")
        else:
            summary.phases[phase] = run_prompt_prune(
                cfg.outputs_dir, review_store=review_store, dry_run=dry_run
            )

        phase = "eligibility"
        eligibility = build_eligibility(cfg, review_store, stock_path)
        summary.phases[phase] = {
            "eligible_count": len(eligibility.eligible),
            "blocked_local_count": len(eligibility.blocked_local),
            "blocked_stock_count": len(eligibility.blocked_stock),
        }
        if not eligibility.eligible:
            raise RuntimeError("No eligible approved SKUs with Stock rows and local assets")

        phase = "uniqueness_checks"
        summary.phases[phase] = run_uniqueness_checks(
            cfg, review_store, stock_path, eligibility.eligible
        )

        phase = "rebuild_xlsx"
        if dry_run:
            summary.phases[phase] = {"dry_run": True, "would_rebuild": str(enriched_path)}
        else:
            row_count = rebuild_stock_enriched(
                cfg,
                review_store,
                stock_path=stock_path,
                products_path=products_path,
                output_path=enriched_path,
            )
            summary.phases[phase] = {"row_count": row_count, "output_path": str(enriched_path)}

        phase = "validate_xlsx"
        if dry_run:
            summary.phases[phase] = {"dry_run": True, "skipped_strict_validation": True}
        else:
            summary.phases[phase] = validate_xlsx_for_eligible(
                cfg, review_store, enriched_path, eligibility.eligible
            )

        if not skip_upload:
            phase = "recreate_products"
            summary.phases[phase] = recreate_approved_products(
                cfg,
                client,
                review_store,
                title_store,
                stock_path,
                eligibility.eligible,
                dry_run=dry_run,
            )
        else:
            log.info("Skipping recreate/upload phase")

        phase = "final_audit"
        if dry_run or skip_upload:
            summary.phases[phase] = {"dry_run": dry_run, "skipped": skip_upload}
        else:
            summary.phases[phase] = run_final_audit(
                cfg, review_store, eligibility.eligible, client=client
            )
            if not summary.phases[phase].get("passed"):
                raise RuntimeError("Final audit failed — see outputs/rebuild_executor/final_audit.json")

        summary.success = True
    except Exception as e:
        summary.stopped_at = phase
        summary.error = str(e)
        log.error("Executor stopped at %s: %s", phase, e)
        save_phase_report(cfg.outputs_dir, "executor_summary", summary.__dict__)
        raise

    save_phase_report(cfg.outputs_dir, "executor_summary", summary.__dict__)
    return summary
