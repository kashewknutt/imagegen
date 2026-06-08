#!/usr/bin/env python3
"""
Update Shopify variant prices (and cost/weight/qty) from Stock.xlsx for every SKU.

Stock.xlsx sell price is often an Excel formula (Labour+rate)*weight — src/stock_pricing
computes the same value when price_2 is not stored as a cached number.

Usage:
  python update_shopify_prices.py --dry-run
  python update_shopify_prices.py
  python update_shopify_prices.py --sku DIAEFHW26001
"""
from __future__ import annotations

import argparse
import logging
import time
from pathlib import Path

from dotenv import load_dotenv

from dedupe_titles_and_upload import _load_shopify_client
from src.config import load_config
from src.stock_pricing import price_fields_from_row
from src.xlsx_ingest import index_by_sku, iter_rows

log = logging.getLogger("update_prices")


def _parse_float(value: object) -> float | None:
    try:
        s = str(value).strip().replace(",", "")
        return float(s) if s else None
    except Exception:
        return None


def _stock_index(stock_path: Path) -> dict[str, dict]:
    rows = index_by_sku(iter_rows(stock_path, ["Total"]), sku_column="SKU")
    return {sku: dict(row.values) for sku, row in rows.items()}


def _enriched_fallback_index(enriched_path: Path) -> dict[str, dict]:
    """Fallback Labour/rate/weight from stock_enriched for supplementary rows."""
    from openpyxl import load_workbook

    if not enriched_path.is_file():
        return {}
    wb = load_workbook(enriched_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        sku = str(vals.get("SKU") or "").strip()
        if sku and sku not in out:
            out[sku] = vals
    wb.close()
    return out


def update_all_prices(
    *,
    cfg,
    client,
    stock_path: Path,
    enriched_path: Path,
    skus: list[str] | None = None,
    dry_run: bool = False,
    sleep_seconds: float = 0.25,
) -> dict:
    stock_by_sku = _stock_index(stock_path)
    enriched_by_sku = _enriched_fallback_index(enriched_path)
    variants = client.fetch_all_variants()
    if skus:
        allow = set(skus)
        variants = [v for v in variants if v["sku"] in allow]

    ok = 0
    skipped = 0
    failed: dict[str, str] = {}
    no_price: list[str] = []

    log.info("Updating prices for %d Shopify variant(s)", len(variants))
    for i, var in enumerate(variants, start=1):
        sku = var["sku"]
        row = stock_by_sku.get(sku) or enriched_by_sku.get(sku) or {}
        price_sell, price_cost, weight_g, qty = price_fields_from_row(row)
        if not price_sell:
            no_price.append(sku)
            skipped += 1
            continue

        variant_id = var.get("variant_id_int")
        if not variant_id:
            failed[sku] = "missing variant_id"
            continue

        current = str(var.get("current_price") or "0")
        log.info(
            "[%d/%d] %s price %s -> %s",
            i,
            len(variants),
            sku,
            current,
            price_sell,
        )
        if dry_run:
            ok += 1
            continue

        try:
            client.rest_variant_update(variant_id=variant_id, sku=sku, price=price_sell)
            if weight_g is not None:
                client.rest_variant_weight(variant_id=variant_id, weight_kg=float(weight_g) / 1000.0)
            inv_id = var.get("inventory_item_id")
            if inv_id:
                cost_f = _parse_float(price_cost)
                if cost_f is not None:
                    client.rest_inventory_item_cost(inventory_item_id=inv_id, cost=cost_f)
                if qty:
                    locs = client.rest_locations()
                    if locs:
                        location_id = int((locs[0] or {}).get("id") or 0)
                        if location_id:
                            client.rest_inventory_set(
                                location_id=location_id,
                                inventory_item_id=inv_id,
                                available=int(qty),
                            )
            ok += 1
        except Exception as e:
            failed[sku] = str(e)
            log.error("[%s] Price update failed: %s", sku, e)
        time.sleep(sleep_seconds)

    report = {
        "dry_run": dry_run,
        "variant_count": len(variants),
        "updated": ok,
        "skipped_no_price": len(no_price),
        "failed": failed,
        "no_price_skus": no_price[:50],
    }
    out = cfg.outputs_dir / "rebuild_executor" / "update_prices.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    import json

    out.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report


def main() -> int:
    load_dotenv()
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(description="Update Shopify prices from Stock.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sku", action="append", help="Limit to SKU(s)")
    parser.add_argument("--sleep", type=float, default=0.25)
    parser.add_argument("--config", default="config.yaml")
    args = parser.parse_args()

    cfg = load_config(args.config)
    client = _load_shopify_client(cfg.outputs_dir)
    report = update_all_prices(
        cfg=cfg,
        client=client,
        stock_path=cfg.xlsx_path,
        enriched_path=cfg.outputs_dir / "stock_enriched.xlsx",
        skus=args.sku,
        dry_run=args.dry_run,
        sleep_seconds=args.sleep,
    )
    log.info(
        "Done: %d updated, %d skipped (no price data), %d failed",
        report["updated"],
        report["skipped_no_price"],
        len(report["failed"]),
    )
    if report["skipped_no_price"]:
        log.warning("%d SKU(s) have no computable price in Stock/enriched", report["skipped_no_price"])
    return 1 if report["failed"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
