#!/usr/bin/env python3
"""
Build an enriched Stock.xlsx export in outputs/ with extra Shopify/product columns.

Usage:
  python build_stock_export.py
  python build_stock_export.py --output outputs/stock_enriched.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from src.media_workspace import resolve_thumbnail_path, thumbnail_relative_path
from src.sku_aliases import canonical_sku
from src.xlsx_ingest import iter_rows

PRODUCT_IMAGE_MAX_PX = 120
PRODUCT_IMAGE_ROW_HEIGHT = 95
PRODUCT_IMAGE_COL_WIDTH = 18


def _norm_sku(value: object) -> str:
    return str(value or "").strip()


def _normalize_category(value: str) -> str:
    v = (value or "").strip()
    if v.lower() == "pandent":
        return "pendant"
    return v


def _load_product_titles(products_path: Path) -> dict[str, str]:
    """Map Variant SKU -> Title from products export."""
    title_by_sku: dict[str, str] = {}
    rows = iter_rows(products_path, ["products_export_1"])
    for row in rows:
        sku = _norm_sku(row.values.get("Variant SKU"))
        if not sku:
            continue
        title = str(row.values.get("Title") or "").strip()
        if title and sku not in title_by_sku:
            title_by_sku[sku] = title
    return title_by_sku


def _thumbnail_image_name(*, category: str, sku: str) -> str:
    cat = _normalize_category(category)
    parts = ["ZOCI"]
    if cat:
        parts.append(cat)
    if sku:
        parts.append(sku)
    return " ".join(parts)


def _thumbnail_image_path(
    *,
    outputs_dir: Path,
    sku: str,
    images_dir: Path | None = None,
    prompt2_version: int | None = None,
) -> Path | None:
    """prompt2 thumbnail (approved/latest), else raw fallback for embedded cells."""
    if not sku:
        return None
    return resolve_thumbnail_path(
        outputs_dir=outputs_dir,
        sku=sku,
        images_dir=images_dir,
        prompt2_version=prompt2_version,
    )


def _product_name_for_sku(*, sku: str, overrides: dict[str, str], title_by_sku: dict[str, str]) -> str:
    for candidate in (sku, canonical_sku(sku)):
        if overrides.get(candidate):
            return overrides[candidate]
        if title_by_sku.get(candidate):
            return title_by_sku[candidate]
    return ""


def _supplementary_rows(
    *,
    stock_columns: list[str],
    stock_skus: set[str],
    overrides: dict[str, str],
    outputs_dir: Path,
    images_dir: Path | None,
    prompt2_versions: dict[str, int | None] | None = None,
) -> list[dict[str, object]]:
    """Rows for titled SKUs with images that are missing from Stock.xlsx."""
    extra: list[dict[str, object]] = []
    for sku in sorted(overrides):
        if not overrides.get(sku) or sku in stock_skus:
            continue
        p2v = (prompt2_versions or {}).get(sku)
        if not _thumbnail_image_path(
            outputs_dir=outputs_dir,
            sku=sku,
            images_dir=images_dir,
            prompt2_version=p2v,
        ):
            continue
        row = {col: "" for col in stock_columns}
        row["SKU"] = sku
        extra.append(row)
    return extra


def _embed_thumbnail_images(
    ws,
    *,
    outputs_dir: Path,
    images_dir: Path | None,
    sku_col_idx: int,
    image_col_idx: int,
    prompt2_versions: dict[str, int | None] | None = None,
    start_row: int = 2,
) -> int:
    """Embed thumbnails in thumbnailImage column cells. Returns count embedded."""
    col_letter = get_column_letter(image_col_idx)
    ws.column_dimensions[col_letter].width = PRODUCT_IMAGE_COL_WIDTH
    embedded = 0
    for row_idx in range(start_row, ws.max_row + 1):
        sku = _norm_sku(ws.cell(row=row_idx, column=sku_col_idx).value)
        p2v = (prompt2_versions or {}).get(sku)
        path = _thumbnail_image_path(
            outputs_dir=outputs_dir,
            sku=sku,
            images_dir=images_dir,
            prompt2_version=p2v,
        )
        if path is None:
            continue
        img = XLImage(str(path))
        w, h = float(img.width), float(img.height)
        if w > 0 and h > 0:
            scale = min(PRODUCT_IMAGE_MAX_PX / w, PRODUCT_IMAGE_MAX_PX / h, 1.0)
            img.width = int(w * scale)
            img.height = int(h * scale)
        ws.add_image(img, f"{col_letter}{row_idx}")
        ws.row_dimensions[row_idx].height = PRODUCT_IMAGE_ROW_HEIGHT
        ws.cell(row=row_idx, column=image_col_idx).value = None
        embedded += 1
    return embedded


def build_export(
    *,
    stock_path: Path,
    products_path: Path,
    outputs_dir: Path,
    output_path: Path,
    stock_sheets: list[str] | None = None,
    product_names: dict[str, str] | None = None,
    images_dir: Path | None = None,
    prompt2_versions: dict[str, int | None] | None = None,
) -> int:
    sheets = stock_sheets or ["Total"]
    stock_rows = iter_rows(stock_path, sheets)
    if not stock_rows:
        raise RuntimeError(f"No rows found in {stock_path} sheets={sheets}")

    title_by_sku = _load_product_titles(products_path)
    overrides = product_names or {}

    stock_columns = list(stock_rows[0].values.keys())
    extra_columns = [
        "productName",
        "thumbnailImage",
        "thumbnail image path",
        "thumbnailImageName",
        "productDescription",
        "hashtag/keyword",
    ]
    headers = stock_columns + extra_columns

    wb = Workbook()
    ws = wb.active
    ws.title = "Total"
    ws.append(headers)

    sku_col_idx = headers.index("SKU") + 1
    image_col_idx = headers.index("thumbnailImage") + 1
    thumb_path_col_idx = headers.index("thumbnail image path") + 1
    stock_skus = {_norm_sku(r.values.get("SKU")) for r in stock_rows if _norm_sku(r.values.get("SKU"))}

    all_rows: list[tuple[dict[str, object], str]] = []
    for row in stock_rows:
        all_rows.append((row.values, str(row.values.get("category") or "").strip()))
    for vals in _supplementary_rows(
        stock_columns=stock_columns,
        stock_skus=stock_skus,
        overrides=overrides,
        outputs_dir=outputs_dir,
        images_dir=images_dir,
        prompt2_versions=prompt2_versions,
    ):
        cat = str(vals.get("category") or "").strip()
        if not cat:
            title = _product_name_for_sku(sku=_norm_sku(vals.get("SKU")), overrides=overrides, title_by_sku=title_by_sku)
            if "bracelet" in title.lower():
                cat = "bracelets"
            elif "pendant" in title.lower() or "necklace" in title.lower():
                cat = "necklaces"
            elif "ring" in title.lower():
                cat = "rings"
            elif "earring" in title.lower() or "stud" in title.lower() or "hoop" in title.lower():
                cat = "earrings"
        all_rows.append((vals, cat))

    for vals, category in all_rows:
        sku = _norm_sku(vals.get("SKU"))

        base = [vals.get(col, "") for col in stock_columns]
        p2v = (prompt2_versions or {}).get(sku)
        thumb_rel = thumbnail_relative_path(outputs_dir=outputs_dir, sku=sku, prompt2_version=p2v)
        extras = [
            _product_name_for_sku(sku=sku, overrides=overrides, title_by_sku=title_by_sku),
            "",  # image embedded after save prep, not a path string
            thumb_rel,
            _thumbnail_image_name(category=category, sku=sku),
            "",
            "",
        ]
        ws.append(base + extras)

    _embed_thumbnail_images(
        ws,
        outputs_dir=outputs_dir,
        images_dir=images_dir,
        sku_col_idx=sku_col_idx,
        image_col_idx=image_col_idx,
        prompt2_versions=prompt2_versions,
    )

    for row_idx in range(2, ws.max_row + 1):
        sku = _norm_sku(ws.cell(row=row_idx, column=sku_col_idx).value)
        if not sku:
            continue
        p2v = (prompt2_versions or {}).get(sku)
        thumb_rel = thumbnail_relative_path(outputs_dir=outputs_dir, sku=sku, prompt2_version=p2v)
        if thumb_rel:
            ws.cell(row=row_idx, column=thumb_path_col_idx).value = thumb_rel

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(all_rows)


def main() -> None:
    root = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Build enriched Stock.xlsx in outputs/")
    parser.add_argument("--stock", type=Path, default=root / "Stock.xlsx")
    parser.add_argument("--products", type=Path, default=root / "products_export_1.xlsx")
    parser.add_argument("--outputs-dir", type=Path, default=root / "outputs")
    parser.add_argument("--output", type=Path, default=root / "outputs" / "stock_enriched.xlsx")
    parser.add_argument(
        "--sheet",
        action="append",
        default=None,
        help="Stock sheet to include (default: Total). Repeat for multiple sheets.",
    )
    args = parser.parse_args()

    from src.config import load_config

    cfg = load_config()
    names_path = cfg.outputs_dir / "title_gen_state.json"
    product_names: dict[str, str] = {}
    if names_path.is_file():
        from dedupe_titles_and_upload import product_names_from_store
        from src.title_store import TitleStore

        product_names = product_names_from_store(TitleStore(names_path))

    review_path = cfg.outputs_dir / "review_state.json"
    prompt2_versions: dict[str, int | None] = {}
    if review_path.is_file():
        from src.review_store import ReviewStore

        review_store = ReviewStore(review_path)
        for sku, rec in review_store.all_records().items():
            approved_title = str(rec.get("title") or "").strip()
            if approved_title:
                product_names[sku] = approved_title
            p2 = rec.get("approved_prompt2_version")
            if p2 is not None:
                prompt2_versions[sku] = int(p2)

    count = build_export(
        stock_path=args.stock,
        products_path=args.products,
        outputs_dir=args.outputs_dir,
        output_path=args.output,
        stock_sheets=args.sheet,
        product_names=product_names or None,
        images_dir=cfg.images_dir,
        prompt2_versions=prompt2_versions or None,
    )
    print(f"Wrote {count} rows to {args.output}")


if __name__ == "__main__":
    main()
