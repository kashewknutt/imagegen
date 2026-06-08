#!/usr/bin/env python3
"""
Audit local media workspaces, Shopify products, and stock_enriched.xlsx alignment.

Usage:
  python audit_workspace.py
  python audit_workspace.py --sku DIAESTR26057
"""
from __future__ import annotations

import argparse
import logging
import re
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

from src.config import load_config
from src.media_workspace import (
    index_sku_media,
    manifest_path,
    sku_workspace_dir,
    thumbnail_relative_path,
)
from src.review_store import ReviewStore
from src.shopify_media_sync import images_for_sku, media_paths_for_sku
from src.shopify_product_dedup import is_active_shopify_product, prefer_canonical_product, primary_sku_from_product
from src.title_store import TitleStore

log = logging.getLogger("audit_workspace")


def _load_xlsx_maps(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    """Return sku -> productName, sku -> thumbnail image path from enriched export."""
    names: dict[str, str] = {}
    paths: dict[str, str] = {}
    if not path.is_file():
        return names, paths
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        sku_i = headers.index("SKU")
        name_i = headers.index("productName")
        path_i = headers.index("thumbnail image path")
    except ValueError:
        wb.close()
        return names, paths
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[sku_i] or "").strip()
        if not sku:
            continue
        names[sku] = str(row[name_i] or "").strip()
        paths[sku] = str(row[path_i] or "").strip()
    wb.close()
    return names, paths


def audit_sku(
    *,
    cfg,
    sku: str,
    shopify_product: dict | None,
    review_store: ReviewStore,
    xlsx_names: dict[str, str],
    xlsx_paths: dict[str, str],
    strict_shopify: bool = False,
) -> tuple[list[str], list[str]]:
    """Return (errors, warnings). Errors fail the audit; warnings are informational."""
    errors: list[str] = []
    warnings: list[str] = []
    idx = index_sku_media(outputs_dir=cfg.outputs_dir, sku=sku)
    ws_dir = sku_workspace_dir(cfg.outputs_dir, sku)

    if not ws_dir.is_dir():
        errors.append("missing outputs workspace folder")
    if not idx.prompt2_versions:
        errors.append("missing prompt2 generated image")
    if not idx.prompt1_versions:
        errors.append("missing prompt1 generated image")
    if not idx.raw_images:
        errors.append("missing raw images in workspace")
    elif not manifest_path(ws_dir).is_file():
        warnings.append("missing media_manifest.json")

    rec = review_store.get_record(sku)
    review_status = str(rec.get("review_status") or "pending_review")
    approved_title = str(rec.get("title") or "").strip()
    p2v = rec.get("approved_prompt2_version")
    expected_thumb = thumbnail_relative_path(
        outputs_dir=cfg.outputs_dir,
        sku=sku,
        prompt2_version=int(p2v) if p2v is not None else None,
    )

    if shopify_product:
        shop_title = str(shopify_product.get("title") or "").strip()
        if approved_title and shop_title and approved_title != shop_title and review_status == "uploaded":
            errors.append(f"Shopify title mismatch: '{shop_title}' vs approved '{approved_title}'")

        expected_images = images_for_sku(cfg, sku, review_store=review_store)
        paths = media_paths_for_sku(cfg, sku, review_store=review_store)
        actual_images = [
            m
            for m in (shopify_product.get("media") or [])
            if str(m.get("content_type") or "IMAGE").upper() != "VIDEO"
        ]
        expected_image_count = len(expected_images)
        expected_video_count = len(paths.get("videos") or [])
        actual_image_count = len(actual_images)

        if actual_image_count < expected_image_count:
            msg = (
                f"Shopify images {actual_image_count}/{expected_image_count} "
                f"(+{expected_video_count} local video(s) not in image count)"
            )
            if review_status == "uploaded" or strict_shopify:
                errors.append(msg)
            else:
                warnings.append(f"needs Shopify sync: {msg}")

        featured = str(shopify_product.get("featured_image_url") or "").lower()
        if featured and "prompt2" not in featured and idx.prompt2_versions:
            thumb_msg = "Shopify featured image is not prompt2"
            if review_status == "uploaded" or strict_shopify:
                errors.append(thumb_msg)
            else:
                warnings.append(thumb_msg)
    elif strict_shopify:
        errors.append("SKU not found on Shopify")
    else:
        warnings.append("SKU not found on Shopify")

    if review_status in {"approved", "uploaded"}:
        xlsx_name = xlsx_names.get(sku, "")
        if approved_title and xlsx_name and approved_title != xlsx_name:
            errors.append(f"XLSX productName mismatch: '{xlsx_name}' vs approved '{approved_title}'")
        xlsx_path = xlsx_paths.get(sku, "")
        if expected_thumb and xlsx_path and expected_thumb != xlsx_path:
            errors.append(f"XLSX thumbnail path mismatch: '{xlsx_path}' vs expected '{expected_thumb}'")
        if approved_title and not xlsx_name:
            errors.append("approved title missing from XLSX productName")

    return errors, warnings


def main() -> int:
    load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")

    parser = argparse.ArgumentParser(description="Audit workspace, Shopify, and XLSX alignment")
    parser.add_argument("--sku", action="append", help="Limit to SKU(s)")
    parser.add_argument(
        "--strict-shopify",
        action="store_true",
        help="Treat Shopify sync gaps and missing products as errors (default: warnings)",
    )
    parser.add_argument(
        "--dedupe-first",
        action="store_true",
        help="Remove duplicate raw/video files in workspaces before auditing",
    )
    args = parser.parse_args()

    cfg = load_config()
    if args.dedupe_first:
        from src.media_workspace import dedupe_all_workspaces

        removed = dedupe_all_workspaces(cfg.outputs_dir)
        total_raw = sum(r.get("raw_removed", 0) for r in removed)
        total_vid = sum(r.get("videos_removed", 0) for r in removed)
        log.info("Deduped %d SKU(s): removed %d raw, %d video duplicate(s)", len(removed), total_raw, total_vid)

    review_store = ReviewStore(cfg.outputs_dir / "review_state.json")
    title_store = TitleStore(cfg.outputs_dir / "title_gen_state.json")
    xlsx_path = cfg.outputs_dir / "stock_enriched.xlsx"
    xlsx_names, xlsx_paths = _load_xlsx_maps(xlsx_path)

    shopify_by_sku: dict[str, dict] = {}
    try:
        from dedupe_titles_and_upload import _load_shopify_client

        client = _load_shopify_client(cfg.outputs_dir)
        after = None
        for _ in range(200):
            page = client.list_products(first=50, after=after, query=None)
            for p in page.get("products") or []:
                if not is_active_shopify_product(p):
                    continue
                sku = primary_sku_from_product(p)
                if not sku:
                    continue
                if sku not in shopify_by_sku:
                    shopify_by_sku[sku] = p
                else:
                    shopify_by_sku[sku] = prefer_canonical_product(shopify_by_sku[sku], p, sku=sku)
            pi = page.get("pageInfo") or {}
            if not pi.get("hasNextPage"):
                break
            after = pi.get("endCursor")
    except Exception as e:
        log.warning("Shopify audit skipped: %s", e)

    def _looks_like_sku(name: str) -> bool:
        n = (name or "").strip()
        if not n or n.startswith("_"):
            return False
        if n.lower() in {"stock", "export"}:
            return False
        if n.startswith("export("):
            return False
        return bool(re.match(r"^[A-Z]{2,}[A-Z0-9_-]+$", n))

    skus: set[str] = set()
    if args.sku:
        skus.update(args.sku)
    else:
        for p in cfg.outputs_dir.iterdir():
            if p.is_dir() and _looks_like_sku(p.name):
                skus.add(p.name)
        for key in title_store.all_records():
            k = str(key).strip()
            if _looks_like_sku(k):
                skus.add(k)

    ok = 0
    bad = 0
    warn_only = 0
    for sku in sorted(skus):
        if not sku:
            continue
        errors, warnings = audit_sku(
            cfg=cfg,
            sku=sku,
            shopify_product=shopify_by_sku.get(sku),
            review_store=review_store,
            xlsx_names=xlsx_names,
            xlsx_paths=xlsx_paths,
            strict_shopify=args.strict_shopify,
        )
        if errors:
            bad += 1
            log.warning("%s — %d error(s): %s", sku, len(errors), "; ".join(errors))
        elif warnings:
            warn_only += 1
            log.info("%s — OK (warnings: %s)", sku, "; ".join(warnings))
        else:
            ok += 1
            log.info("%s — OK", sku)

    log.info(
        "Audit complete: %d OK, %d warnings-only, %d errors (of %d SKU(s))",
        ok,
        warn_only,
        bad,
        len(skus),
    )
    return 1 if bad else 0


if __name__ == "__main__":
    raise SystemExit(main())
